"""
Breakdown Bulk Import Script
============================
Reads the Excel breakdown file and registers each row as a breakdown in the system.

Usage:
    python3 import_breakdowns.py \
        --file  "2026 - BREAK DOWN-1.xlsx" \
        --sheet "2026 Downtime" \
        --server https://www.inspet.pro \
        --user  ADMIN-001 \
        --password "your_password"

    Optional flags:
        --dry-run        Print what would be registered without actually sending
        --skip-planned   Skip rows where Condition = PLANNED
        --date-from      Only import from this date  e.g. 2026-05-01
        --date-to        Only import up to this date e.g. 2026-05-06
"""

import argparse
import sys
import time
from datetime import datetime, date

try:
    import openpyxl
    import requests
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip3 install openpyxl requests")
    sys.exit(1)


# ── Column name detection helpers ────────────────────────────────────────────

KNOWN_HEADERS = {
    "production_date": ["production date", "prod date", "date", "production_date"],
    "machine_no":      ["mach. no.", "machine no", "mach no", "machine_no", "mach.no.", "machine number", "mach no."],
    "from_time":       ["from", "start", "from time", "start time"],
    "to_time":         ["to", "end", "to time", "end time"],
    "reason":          ["reason", "breakdown reason", "cause", "description"],
    "condition":       ["condition", "type", "planned/unplanned"],
    "shift":           ["shift", "shift day/night", "day/night"],
}


def find_header_row(ws):
    """Find the row that contains column headers."""
    for row_idx in range(1, 20):
        row_vals = [str(ws.cell(row=row_idx, column=c).value or "").lower().strip()
                    for c in range(1, ws.max_column + 1)]
        # Header row likely contains "machine" or "reason" or "from"
        if any(k in " ".join(row_vals) for k in ["machine", "reason", "from"]):
            return row_idx
    return 1


def map_columns(ws, header_row):
    """Map field names to column indices."""
    headers = {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=c).value
    }
    col = {}
    for field, candidates in KNOWN_HEADERS.items():
        for hdr_cell, col_idx in headers.items():
            if str(hdr_cell).lower().strip() in candidates:
                col[field] = col_idx
                break
    return col, headers


def parse_date(val):
    """Parse a date value from Excel (datetime, date, or string)."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_time(val):
    """Parse a time value from Excel (datetime, time, or string like 4:00 AM)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if hasattr(val, "hour"):   # time object
        return val.strftime("%H:%M")
    # Numeric (Excel time fraction 0–1)
    if isinstance(val, float):
        total_min = round(val * 24 * 60)
        return f"{total_min // 60:02d}:{total_min % 60:02d}"
    s = str(val).strip()
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return None


def to_iso(date_str, time_str):
    """Combine date + time into ISO 8601 datetime string."""
    if not date_str or not time_str:
        return None
    return f"{date_str}T{time_str}:00"


# ── API helpers ───────────────────────────────────────────────────────────────

def login(server, employee_number, password):
    url = f"{server.rstrip('/')}/api/auth/login"
    r = requests.post(url, json={"employee_number": employee_number, "password": password}, timeout=15)
    r.raise_for_status()
    return r.json()["access_token"]


def get_machines(server, token):
    url = f"{server.rstrip('/')}/api/machines"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=15)
    r.raise_for_status()
    return {str(m["number"]).strip(): m["id"] for m in r.json()}


def post_breakdown(server, token, payload):
    url = f"{server.rstrip('/')}/api/breakdowns"
    r = requests.post(url, json=payload, headers={"Authorization": f"Bearer {token}"}, timeout=15)
    return r


# ── Main import logic ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Bulk import breakdowns from Excel")
    parser.add_argument("--file",          required=True,  help="Path to Excel file")
    parser.add_argument("--sheet",         default=None,   help="Sheet name (default: first sheet)")
    parser.add_argument("--server",        required=True,  help="Server URL e.g. https://www.inspet.pro")
    parser.add_argument("--user",          required=True,  help="Admin employee number")
    parser.add_argument("--password",      required=True,  help="Admin password")
    parser.add_argument("--dry-run",       action="store_true", help="Print rows without sending")
    parser.add_argument("--skip-planned",  action="store_true", help="Skip PLANNED condition rows")
    parser.add_argument("--date-from",     default=None,   help="Import from date YYYY-MM-DD")
    parser.add_argument("--date-to",       default=None,   help="Import up to date YYYY-MM-DD")
    args = parser.parse_args()

    # Load workbook
    print(f"Loading: {args.file}")
    wb = openpyxl.load_workbook(args.file, data_only=True)
    if args.sheet:
        ws = wb[args.sheet]
    else:
        # Try common sheet names
        for name in ["2026 Downtime", "2026 BREAKDOWN CODE", "Downtime", "Sheet1"]:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        else:
            ws = wb.active
    print(f"Sheet: {ws.title}  |  Rows: {ws.max_row}")

    # Find headers
    hdr_row = find_header_row(ws)
    col, all_headers = map_columns(ws, hdr_row)
    print(f"Header row: {hdr_row}")
    print(f"Column mapping: { {k: v for k,v in col.items()} }")

    required = ["production_date", "machine_no", "from_time", "to_time", "reason"]
    missing  = [r for r in required if r not in col]
    if missing:
        print(f"\n❌ Cannot find columns: {missing}")
        print(f"   Available headers: {list(all_headers.keys())}")
        sys.exit(1)

    # Login
    machines_by_num = {}
    if not args.dry_run:
        print(f"\nConnecting to {args.server} ...")
        token = login(args.server, args.user, args.password)
        print("✓ Logged in")
        machines_by_num = get_machines(args.server, token)
        print(f"✓ Found {len(machines_by_num)} machines in system")

    # Date filter
    df = datetime.strptime(args.date_from, "%Y-%m-%d").date() if args.date_from else None
    dt = datetime.strptime(args.date_to,   "%Y-%m-%d").date() if args.date_to   else None

    # Process rows
    success = skip_planned = skip_no_machine = skip_no_date = skip_no_time = errors = 0
    total_rows = 0

    for row_idx in range(hdr_row + 1, ws.max_row + 1):
        def cell(field):
            return ws.cell(row=row_idx, column=col[field]).value if field in col else None

        raw_date    = cell("production_date")
        raw_mach    = cell("machine_no")
        raw_from    = cell("from_time")
        raw_to      = cell("to_time")
        raw_reason  = cell("reason")
        raw_cond    = str(cell("condition") or "").strip().upper() if "condition" in col else ""

        # Skip empty rows
        if raw_date is None and raw_mach is None:
            continue

        total_rows += 1
        date_str   = parse_date(raw_date)
        from_str   = parse_time(raw_from)
        to_str     = parse_time(raw_to)
        machine_no = str(raw_mach or "").strip()
        reason     = str(raw_reason or "").strip()

        # Skip planned if requested
        if args.skip_planned and "PLANNED" in raw_cond and "UNPLANNED" not in raw_cond:
            skip_planned += 1
            continue

        # Date filter
        if date_str:
            row_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            if df and row_date < df:
                continue
            if dt and row_date > dt:
                continue
        else:
            skip_no_date += 1
            print(f"  Row {row_idx}: ⚠ Cannot parse date '{raw_date}'")
            continue

        if not from_str or not to_str:
            skip_no_time += 1
            print(f"  Row {row_idx}: ⚠ Cannot parse time FROM='{raw_from}' TO='{raw_to}'")
            continue

        start_iso = to_iso(date_str, from_str)
        end_iso   = to_iso(date_str, to_str)

        # Dry run
        if args.dry_run:
            print(f"  Row {row_idx}: Mach={machine_no}  {start_iso} → {end_iso}  | {reason[:50]}")
            success += 1
            continue

        # Resolve machine ID
        machine_id = machines_by_num.get(machine_no)
        if not machine_id:
            # Try without leading zeros
            machine_id = machines_by_num.get(machine_no.lstrip("0"))
        if not machine_id:
            skip_no_machine += 1
            if skip_no_machine <= 10:
                print(f"  Row {row_idx}: ⚠ Machine '{machine_no}' not found in system")
            continue

        payload = {
            "machine_id":        machine_id,
            "start_time":        start_iso,
            "end_time":          end_iso,
            "brief_description": reason[:200],
            "repair_description": "",
        }

        r = post_breakdown(args.server, token, payload)
        if r.status_code in (200, 201):
            success += 1
            if success % 50 == 0:
                print(f"  ... {success} registered so far")
        else:
            errors += 1
            if errors <= 5:
                print(f"  Row {row_idx}: ❌ HTTP {r.status_code} — {r.text[:100]}")

        time.sleep(0.05)   # 50ms between requests to avoid rate limiting

    # Summary
    print(f"""
══════════════════════════════════
Import complete
  Total data rows   : {total_rows}
  ✓ Registered      : {success}
  ⚠ No machine match: {skip_no_machine}
  ⚠ Skipped planned : {skip_planned}
  ⚠ Bad date        : {skip_no_date}
  ⚠ Bad time        : {skip_no_time}
  ❌ Errors          : {errors}
══════════════════════════════════
""")


if __name__ == "__main__":
    main()
