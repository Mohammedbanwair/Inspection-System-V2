from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
MDB_UPLOADS_DIR = UPLOADS_DIR / "mdb"
MDB_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
load_dotenv(ROOT_DIR / '.env')

import os
import re
import io
import uuid
import time
import logging
import bcrypt
import jwt
import pymongo.errors
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.gzip import GZipMiddleware
from pydantic import BaseModel, EmailStr
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
import ssl
import certifi
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

mongo_url = os.environ['MONGO_URL']
ssl_context = ssl.create_default_context(cafile=certifi.where())
ssl_context.minimum_version = ssl.TLSVersion.TLSv1_2
client = AsyncIOMotorClient(
    mongo_url,
    tls=True,
    tlsCAFile=certifi.where(),
    serverSelectionTimeoutMS=30000,
)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGO = "HS256"
ACCESS_TOKEN_MIN = int(os.environ.get("ACCESS_TOKEN_MINUTES", 60 * 24))

app = FastAPI()
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")
api = APIRouter(prefix="/api")

# --- Rate limiting (in-memory, per IP) ---
_login_attempts: dict = defaultdict(list)
_RL_WINDOW = 15 * 60   # 15 minutes window
_RL_MAX    = 5          # max failed attempts before block

def _rl_check(ip: str) -> None:
    now = time.time()
    recent = [t for t in _login_attempts[ip] if now - t < _RL_WINDOW]
    if recent:
        _login_attempts[ip] = recent
    else:
        _login_attempts.pop(ip, None)
    if len(_login_attempts.get(ip, [])) >= _RL_MAX:
        raise HTTPException(429, "تم تجاوز عدد المحاولات المسموح بها. حاول مرة أخرى بعد 15 دقيقة.")

def _rl_fail(ip: str) -> None:
    _login_attempts[ip].append(time.time())

def _rl_clear(ip: str) -> None:
    _login_attempts.pop(ip, None)

_reg_attempts: dict = defaultdict(list)
_RL_REG_WINDOW = 60 * 60
_RL_REG_MAX = 5

def _rl_check_reg(ip: str) -> None:
    now = time.time()
    recent = [t for t in _reg_attempts[ip] if now - t < _RL_REG_WINDOW]
    if recent:
        _reg_attempts[ip] = recent
    else:
        _reg_attempts.pop(ip, None)
    if len(_reg_attempts.get(ip, [])) >= _RL_REG_MAX:
        raise HTTPException(429, "تم تجاوز عدد طلبات التسجيل المسموح بها. حاول مرة أخرى بعد ساعة.")

def _rl_fail_reg(ip: str) -> None:
    _reg_attempts[ip].append(time.time())

# Export rate limiting: max 10 exports per minute per IP
_export_attempts: dict = defaultdict(list)
_RL_EXPORT_WINDOW = 60
_RL_EXPORT_MAX = 10

def _rl_check_export(ip: str) -> None:
    now = time.time()
    recent = [t for t in _export_attempts[ip] if now - t < _RL_EXPORT_WINDOW]
    if recent:
        _export_attempts[ip] = recent
    else:
        _export_attempts.pop(ip, None)
    if len(_export_attempts.get(ip, [])) >= _RL_EXPORT_MAX:
        raise HTTPException(429, "تم تجاوز الحد المسموح به لعمليات التصدير. حاول مرة أخرى بعد دقيقة.")
    _export_attempts[ip].append(now)

logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ── Audit log helper ─────────────────────────────────────────────
async def _audit(actor_id: str, actor_name: str, action: str,
                 resource_type: str, resource_id: str = "", details: str = "") -> None:
    try:
        await db.audit_logs.insert_one({
            "id": str(uuid.uuid4()),
            "actor_id": actor_id,
            "actor_name": actor_name,
            "action": action,
            "resource_type": resource_type,
            "resource_id": resource_id,
            "details": details,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass

# ── Notification helper ───────────────────────────────────────────
async def _notify(ntype: str, title_ar: str, body_ar: str, ref_id: str = "") -> None:
    try:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": ntype,
            "title_ar": title_ar,
            "body_ar": body_ar,
            "ref_id": ref_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass

CATS = ["electrical", "mechanical", "chiller", "panels_main", "panels_sub", "cooling_tower", "preventive"]
Category = Literal["electrical", "mechanical", "chiller", "panels_main", "panels_sub", "cooling_tower", "preventive"]
TargetType = Literal["machine", "chiller", "panel", "cooling_tower"]
NO_COOLDOWN_CATS = {"cooling_tower", "preventive"}


def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()


def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False


def create_token(user_id: str, email: str, role: str) -> str:
    payload = {"sub": user_id, "email": email, "role": role,
               "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_MIN)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(401, "User not found")
    return user


async def require_admin(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return user


def allowed_categories(user: dict) -> List[str]:
    if user["role"] == "admin":
        return CATS
    if user["role"] == "helper":
        return ["preventive"]
    sp = user.get("specialty")
    if sp == "electrical":
        return ["electrical", "panels_main", "panels_sub"]
    if sp == "mechanical":
        return ["mechanical", "chiller", "cooling_tower"]
    return []


# Mapping: which target_type each category requires
CAT_TARGET = {
    "electrical": "machine",
    "mechanical": "machine",
    "chiller": "chiller",
    "panels_main": "panel",
    "panels_sub": "panel",
    "cooling_tower": "cooling_tower",
    "preventive": "machine",
}


class LoginIn(BaseModel):
    employee_number: str
    password: str


class UserCreate(BaseModel):
    employee_number: str
    password: str
    name: str
    role: Literal["admin", "technician", "helper"] = "technician"
    specialty: Optional[Literal["electrical", "mechanical"]] = None


class UserUpdate(BaseModel):
    name: Optional[str] = None
    password: Optional[str] = None
    role: Optional[Literal["admin", "technician", "helper"]] = None
    specialty: Optional[Literal["electrical", "mechanical"]] = None


class EntityCreate(BaseModel):
    number: str
    name: Optional[str] = ""
    group: Optional[str] = None


class EntityUpdate(BaseModel):
    number: Optional[str] = None
    name: Optional[str] = None
    group: Optional[str] = None


class QuestionCreate(BaseModel):
    category: Category
    text: str
    order: int = 0
    answer_type: Literal["yes_no", "numeric"] = "yes_no"
    unit: Optional[str] = "°C"
    section: Optional[str] = None


class QuestionUpdate(BaseModel):
    category: Optional[Category] = None
    text: Optional[str] = None
    order: Optional[int] = None
    answer_type: Optional[Literal["yes_no", "numeric"]] = None
    unit: Optional[str] = None
    section: Optional[str] = None


class AnswerIn(BaseModel):
    question_id: str
    answer: Optional[bool] = None
    numeric_value: Optional[float] = None
    note: Optional[str] = ""
    spare_part: Optional[str] = ""


class InspectionCreate(BaseModel):
    category: Category
    target_type: TargetType
    target_id: str
    answers: List[AnswerIn]
    notes: Optional[str] = ""
    time_mc_received: Optional[str] = ""
    time_delivered: Optional[str] = ""


class InspectionUpdate(BaseModel):
    answers: List[AnswerIn]
    notes: Optional[str] = ""
    time_mc_received: Optional[str] = ""
    time_delivered: Optional[str] = ""


class PreventivePlanCreate(BaseModel):
    machine_id: str
    scheduled_date: str  # YYYY-MM-DD


class BreakdownCreate(BaseModel):
    machine_id: str
    brief_description: str
    repair_description: Optional[str] = ""
    start_time: Optional[str] = ""
    end_time: Optional[str] = ""


class DowntimeReasonCreate(BaseModel):
    text: str
    specialty: Optional[Literal["electrical", "mechanical"]] = None

class DowntimeReasonUpdate(BaseModel):
    text: Optional[str] = None
    specialty: Optional[Literal["electrical", "mechanical"]] = None


class RegisterRequest(BaseModel):
    employee_number: str
    name: str
    password: str


# ---------- Auth ----------
@api.post("/auth/register")
async def register_request(body: RegisterRequest, request: Request):
    ip = request.client.host if request.client else "unknown"
    _rl_check_reg(ip)
    emp = body.employee_number.strip().upper()
    if not emp or not body.name.strip() or not body.password:
        _rl_fail_reg(ip)
        raise HTTPException(400, "جميع الحقول مطلوبة")
    # Check no existing user with same employee number
    if await db.users.find_one({"employee_number": emp}):
        raise HTTPException(400, "رقم الموظف مستخدم مسبقاً")
    # Check no pending request with same employee number
    if await db.registration_requests.find_one({"employee_number": emp, "status": "pending"}):
        raise HTTPException(400, "يوجد طلب تسجيل معلق لهذا الرقم")
    doc = {
        "id": str(uuid.uuid4()),
        "employee_number": emp,
        "name": body.name.strip(),
        "password_hash": hash_password(body.password),
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": datetime.now(timezone.utc) + timedelta(days=90),
    }
    await db.registration_requests.insert_one(doc)
    _rl_fail_reg(ip)
    await _notify("new_registration",
                  f"طلب تسجيل جديد",
                  f"طلب {body.name.strip()} ({emp}) الانضمام للنظام",
                  doc["id"])
    doc.pop("_id", None); doc.pop("password_hash", None)
    return doc


@api.get("/registration-requests/count")
async def count_registration_requests(_=Depends(require_admin)):
    n = await db.registration_requests.count_documents({"status": "pending"})
    return {"count": n}


@api.get("/registration-requests")
async def list_registration_requests(_=Depends(require_admin)):
    items = await db.registration_requests.find(
        {"status": "pending"}, {"_id": 0, "password_hash": 0}
    ).sort("created_at", 1).to_list(500)
    return items


class ApproveBody(BaseModel):
    specialty: Literal["electrical", "mechanical"]


@api.post("/registration-requests/{rid}/approve")
async def approve_registration(
    rid: str,
    body: ApproveBody,
    admin=Depends(require_admin),
):
    specialty = body.specialty
    req = await db.registration_requests.find_one({"id": rid, "status": "pending"})
    if not req:
        raise HTTPException(404, "الطلب غير موجود أو تمت معالجته")
    emp = req["employee_number"]
    if await db.users.find_one({"employee_number": emp}):
        raise HTTPException(400, "رقم الموظف مستخدم مسبقاً")
    user_doc = {
        "id": str(uuid.uuid4()),
        "employee_number": emp,
        "name": req["name"],
        "role": "technician",
        "specialty": specialty,
        "password_hash": req["password_hash"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user_doc)
    await db.registration_requests.update_one(
        {"id": rid}, {"$set": {"status": "approved", "specialty": specialty,
                               "approved_at": datetime.now(timezone.utc).isoformat(),
                               "expires_at": datetime.now(timezone.utc) + timedelta(days=30)}}
    )
    await _audit(admin["id"], admin["name"], "approve", "registration", rid,
                 f"{req['name']} ({emp})")
    return {"ok": True}


@api.delete("/registration-requests/{rid}")
async def reject_registration(rid: str, _=Depends(require_admin)):
    res = await db.registration_requests.update_one(
        {"id": rid, "status": "pending"},
        {"$set": {"status": "rejected", "rejected_at": datetime.now(timezone.utc).isoformat(),
                  "expires_at": datetime.now(timezone.utc) + timedelta(days=7)}}
    )
    if res.modified_count == 0:
        raise HTTPException(404, "الطلب غير موجود أو تمت معالجته")
    return {"ok": True}


@api.post("/auth/login")
async def login(body: LoginIn, response: Response, request: Request):
    ip = request.client.host if request.client else "unknown"
    _rl_check(ip)
    emp = body.employee_number.strip().upper()
    user = await db.users.find_one({"employee_number": emp})
    if not user or not verify_password(body.password, user["password_hash"]):
        _rl_fail(ip)
        raise HTTPException(401, "رقم الموظف أو كلمة المرور غير صحيحة")
    _rl_clear(ip)
    token = create_token(user["id"], user["employee_number"], user["role"])
    response.set_cookie(key="access_token", value=token, httponly=True, secure=True,
                        samesite="strict", max_age=ACCESS_TOKEN_MIN * 60, path="/")
    return {"token": token, "user": {
        "id": user["id"], "employee_number": user["employee_number"], "name": user["name"],
        "role": user["role"], "specialty": user.get("specialty"),
    }}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


# ---------- Users ----------
@api.get("/users")
async def list_users(_=Depends(require_admin)):
    return await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)


@api.post("/users")
async def create_user(body: UserCreate, admin=Depends(require_admin)):
    emp = body.employee_number.strip().upper()
    if not emp:
        raise HTTPException(400, "رقم الموظف مطلوب")
    if await db.users.find_one({"employee_number": emp}):
        raise HTTPException(400, "رقم الموظف مستخدم مسبقاً")
    specialty = body.specialty if body.role == "technician" else None
    if body.role == "technician" and not specialty:
        raise HTTPException(400, "يجب تحديد تخصص الفني")
    doc = {"id": str(uuid.uuid4()), "employee_number": emp, "name": body.name,
           "role": body.role, "specialty": specialty,
           "password_hash": hash_password(body.password),
           "created_at": datetime.now(timezone.utc).isoformat()}
    try:
        await db.users.insert_one(doc)
    except pymongo.errors.DuplicateKeyError:
        raise HTTPException(400, "رقم الموظف مستخدم مسبقاً")
    doc.pop("password_hash"); doc.pop("_id", None)
    await _audit(admin["id"], admin["name"], "create", "user", doc["id"], f"{body.name} ({emp})")
    return doc


@api.patch("/users/{uid}")
async def update_user(uid: str, body: UserUpdate, _=Depends(require_admin)):
    upd = {}
    if body.name is not None: upd["name"] = body.name
    if body.role is not None: upd["role"] = body.role
    if body.specialty is not None: upd["specialty"] = body.specialty
    if body.password: upd["password_hash"] = hash_password(body.password)
    if not upd: raise HTTPException(400, "لا توجد حقول للتحديث")
    await db.users.update_one({"id": uid}, {"$set": upd})
    u = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
    if not u: raise HTTPException(404, "غير موجود")
    return u


@api.delete("/users/{uid}")
async def delete_user(uid: str, admin=Depends(require_admin)):
    if uid == admin["id"]:
        raise HTTPException(400, "لا يمكن حذف حسابك")
    target = await db.users.find_one({"id": uid}, {"_id": 0, "name": 1, "employee_number": 1})
    res = await db.users.delete_one({"id": uid})
    if res.deleted_count == 0:
        raise HTTPException(404, "غير موجود")
    if target:
        await _audit(admin["id"], admin["name"], "delete", "user", uid,
                     f"{target.get('name','')} ({target.get('employee_number','')})")
    return {"ok": True}


# ---------- Entity factory (chillers only now) ----------
def make_entity_routes(coll_name: str, label: str, route: str):
    @api.get(route)
    async def list_items(user=Depends(get_current_user)):
        return await db[coll_name].find({}, {"_id": 0}).sort([("sort_order", 1), ("number", 1)]).to_list(2000)

    @api.post(route)
    async def create_item(body: EntityCreate, _=Depends(require_admin)):
        num = body.number.strip()
        if await db[coll_name].find_one({"number": num}):
            raise HTTPException(400, f"{label}: الرقم موجود مسبقاً")
        count = await db[coll_name].count_documents({})
        doc = {"id": str(uuid.uuid4()), "number": num, "name": body.name or "",
               "sort_order": count, "created_at": datetime.now(timezone.utc).isoformat()}
        await db[coll_name].insert_one(doc); doc.pop("_id", None)
        return doc

    @api.patch(route + "/reorder")
    async def reorder_items(body: dict, _=Depends(require_admin)):
        ids = body.get("ids", [])
        for i, iid in enumerate(ids):
            await db[coll_name].update_one({"id": iid}, {"$set": {"sort_order": i}})
        return {"ok": True}

    @api.patch(route + "/{iid}")
    async def patch_item(iid: str, body: EntityUpdate, _=Depends(require_admin)):
        upd = {k: v for k, v in body.model_dump(exclude_none=True).items()}
        if not upd: raise HTTPException(400, "لا توجد حقول")
        await db[coll_name].update_one({"id": iid}, {"$set": upd})
        d = await db[coll_name].find_one({"id": iid}, {"_id": 0})
        if not d: raise HTTPException(404, f"{label} غير موجود")
        return d

    @api.delete(route + "/{iid}")
    async def del_item(iid: str, _=Depends(require_admin)):
        res = await db[coll_name].delete_one({"id": iid})
        if res.deleted_count == 0: raise HTTPException(404, f"{label} غير موجود")
        return {"ok": True}


make_entity_routes("chillers", "الشيلر", "/chillers")
make_entity_routes("cooling_towers", "برج التبريد", "/cooling-towers")


# ---------- Machines (with group A/B and custom sort_order) ----------
class MachineCreate(BaseModel):
    number: str
    name: Optional[str] = ""
    group: Literal["A", "B"] = "A"
    manufacturing_year: Optional[str] = ""
    serial_number: Optional[str] = ""
    power_consumption: Optional[str] = ""

class MachineUpdate(BaseModel):
    number: Optional[str] = None
    name: Optional[str] = None
    group: Optional[Literal["A", "B"]] = None
    manufacturing_year: Optional[str] = None
    serial_number: Optional[str] = None
    power_consumption: Optional[str] = None


@api.get("/machines")
async def list_machines(group: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {}
    if group: q["group"] = group
    items = await db.machines.find(q, {"_id": 0}).sort([("sort_order", 1), ("number", 1)]).to_list(2000)
    return items


@api.post("/machines")
async def create_machine(body: MachineCreate, admin=Depends(require_admin)):
    num = body.number.strip()
    if await db.machines.find_one({"number": num}):
        raise HTTPException(400, "المكينة: الرقم موجود مسبقاً")
    count = await db.machines.count_documents({})
    doc = {"id": str(uuid.uuid4()), "number": num, "name": body.name or "",
           "group": body.group, "sort_order": count,
           "manufacturing_year": body.manufacturing_year or "",
           "serial_number": body.serial_number or "",
           "power_consumption": body.power_consumption or "",
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.machines.insert_one(doc); doc.pop("_id", None)
    await _audit(admin["id"], admin["name"], "create", "machine", doc["id"], f"مكينة {num}")
    return doc


@api.patch("/machines/reorder")
async def reorder_machines(body: dict, _=Depends(require_admin)):
    # body = {"ids": ["id1","id2",...]} ordered list
    ids = body.get("ids", [])
    for i, mid in enumerate(ids):
        await db.machines.update_one({"id": mid}, {"$set": {"sort_order": i}})
    return {"ok": True}


@api.patch("/machines/{iid}")
async def patch_machine(iid: str, body: MachineUpdate, _=Depends(require_admin)):
    upd = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not upd: raise HTTPException(400, "لا توجد حقول")
    await db.machines.update_one({"id": iid}, {"$set": upd})
    d = await db.machines.find_one({"id": iid}, {"_id": 0})
    if not d: raise HTTPException(404, "المكينة غير موجودة")
    return d


@api.get("/machines/export/excel")
async def export_machines_excel(request: Request, _=Depends(require_admin)):
    _rl_check_export(request.client.host)
    machines = await db.machines.find({}, {"_id": 0}).sort([("sort_order", 1), ("number", 1)]).to_list(2000)

    PURPLE       = "6B2D6B"
    PURPLE_LIGHT = "F3E8F3"
    WHITE        = "FFFFFF"
    HEADER_BG    = "4A1442"
    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color=PURPLE)

    HEADERS    = ["#", "Machine Number", "Name / Model", "Group",
                  "Serial Number", "Manufacturing Year", "Power Consumption"]
    COL_WIDTHS = [5,   18,               28,              10,
                  22,                  18,                18]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Machines"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = get_column_letter(len(HEADERS))

    ws.row_dimensions[1].height = 8
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=WHITE)

    ws.row_dimensions[2].height = 80
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].fill = PatternFill("solid", fgColor=WHITE)
    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width  = sum(COL_WIDTHS) * 7
        img.height = 82
        img.anchor = "A2"
        ws.add_image(img)

    ws.row_dimensions[3].height = 38
    ws.merge_cells(f"A3:{last_col}3")
    tc = ws["A3"]
    tc.value = "Machine List"
    tc.font  = Font(name="Arial", bold=True, size=16, color=PURPLE)
    tc.fill  = PatternFill("solid", fgColor=WHITE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = Border(bottom=Side(style="medium", color=PURPLE))

    ws.row_dimensions[4].height = 28
    ws.merge_cells(f"A4:{last_col}4")
    ic = ws["A4"]
    ic.value = f"   Total Machines: {len(machines)}"
    ic.font  = Font(name="Arial", bold=True, size=12, color=WHITE)
    ic.fill  = PatternFill("solid", fgColor=HEADER_BG)
    ic.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[5].height = 30
    for ci, hdr in enumerate(HEADERS, 1):
        cl   = get_column_letter(ci)
        cell = ws[f"{cl}5"]
        cell.value = hdr
        cell.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=thick, bottom=thick,
            left=thick if ci == 1 else thin,
            right=thick if ci == len(HEADERS) else thin,
        )

    for ri, m in enumerate(machines):
        row = 6 + ri
        ws.row_dimensions[row].height = 22
        bg = PURPLE_LIGHT if ri % 2 == 0 else WHITE
        values = [
            ri + 1,
            m.get("number", ""),
            m.get("name", "") or "—",
            f"Group {m.get('group', 'A')}",
            m.get("serial_number", "") or "—",
            m.get("manufacturing_year", "") or "—",
            m.get("power_consumption", "") or "—",
        ]
        for ci, val in enumerate(values, 1):
            cl   = get_column_letter(ci)
            cell = ws[f"{cl}{row}"]
            cell.value = val
            cell.font  = Font(name="Arial", size=10, bold=(ci == 2))
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if ci in (1, 4) else "left", vertical="center")
            cell.border = Border(
                top=thin, bottom=thin,
                left=thick if ci == 1 else thin,
                right=thick if ci == len(HEADERS) else thin,
            )

    last_row = 5 + max(len(machines), 1)
    for ci in range(1, len(HEADERS) + 1):
        cl   = get_column_letter(ci)
        cell = ws[f"{cl}{last_row}"]
        b = cell.border
        cell.border = Border(top=b.top, bottom=thick, left=b.left, right=b.right)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="machines.xlsx"'},
    )


@api.delete("/machines/{iid}")
async def delete_machine(iid: str, admin=Depends(require_admin)):
    m = await db.machines.find_one({"id": iid}, {"_id": 0, "number": 1})
    res = await db.machines.delete_one({"id": iid})
    if res.deleted_count == 0: raise HTTPException(404, "المكينة غير موجودة")
    await _audit(admin["id"], admin["name"], "delete", "machine", iid, f"مكينة {m.get('number','') if m else ''}")
    return {"ok": True}


# ---------- Panels (with panel_type: main/sub) ----------
class PanelCreate(BaseModel):
    number: str
    name: Optional[str] = ""
    panel_type: Literal["main", "sub"] = "sub"

class PanelUpdate(BaseModel):
    number: Optional[str] = None
    name: Optional[str] = None
    panel_type: Optional[Literal["main", "sub"]] = None


@api.get("/panels")
async def list_panels(panel_type: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {}
    if panel_type: q["panel_type"] = panel_type
    return await db.panels.find(q, {"_id": 0}).sort([("sort_order", 1), ("number", 1)]).to_list(2000)


@api.post("/panels")
async def create_panel(body: PanelCreate, _=Depends(require_admin)):
    num = body.number.strip()
    if await db.panels.find_one({"number": num}):
        raise HTTPException(400, "اللوحة: الرقم موجود مسبقاً")
    count = await db.panels.count_documents({})
    doc = {"id": str(uuid.uuid4()), "number": num, "name": body.name or "",
           "panel_type": body.panel_type, "sort_order": count,
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.panels.insert_one(doc); doc.pop("_id", None)
    return doc


@api.patch("/panels/reorder")
async def reorder_panels(body: dict, _=Depends(require_admin)):
    ids = body.get("ids", [])
    for i, iid in enumerate(ids):
        await db.panels.update_one({"id": iid}, {"$set": {"sort_order": i}})
    return {"ok": True}


@api.patch("/panels/{iid}")
async def patch_panel(iid: str, body: PanelUpdate, _=Depends(require_admin)):
    upd = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not upd: raise HTTPException(400, "لا توجد حقول")
    await db.panels.update_one({"id": iid}, {"$set": upd})
    d = await db.panels.find_one({"id": iid}, {"_id": 0})
    if not d: raise HTTPException(404, "اللوحة غير موجودة")
    return d


@api.delete("/panels/{iid}")
async def delete_panel(iid: str, _=Depends(require_admin)):
    res = await db.panels.delete_one({"id": iid})
    if res.deleted_count == 0: raise HTTPException(404, "اللوحة غير موجودة")
    return {"ok": True}


# ---------- Questions ----------
@api.get("/questions")
async def list_questions(category: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {}
    if user["role"] != "admin":
        allowed = allowed_categories(user)
        if category:
            if category not in allowed:
                return []
            q["category"] = category
        else:
            q["category"] = {"$in": allowed}
    elif category:
        q["category"] = category
    return await db.questions.find(q, {"_id": 0}).sort([("category", 1), ("order", 1)]).to_list(2000)


@api.post("/questions")
async def create_question(body: QuestionCreate, admin=Depends(require_admin)):
    numeric_allowed = ("panels_main", "panels_sub", "chiller")
    answer_type = body.answer_type if body.category in numeric_allowed else "yes_no"
    doc = {"id": str(uuid.uuid4()), "category": body.category,
           "text": body.text, "order": body.order,
           "answer_type": answer_type,
           "unit": body.unit if answer_type == "numeric" else None,
           "created_at": datetime.now(timezone.utc).isoformat()}
    if body.section:
        doc["section"] = body.section
    await db.questions.insert_one(doc); doc.pop("_id", None)
    await _audit(admin["id"], admin["name"], "create", "question", doc["id"], body.text[:80])
    return doc


@api.patch("/questions/{qid}")
async def update_question(qid: str, body: QuestionUpdate, _=Depends(require_admin)):
    upd = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    numeric_allowed = ("panels_main", "panels_sub", "chiller")
    target_cat = upd.get("category")
    if target_cat and target_cat not in numeric_allowed:
        upd["answer_type"] = "yes_no"
    if not upd: raise HTTPException(400, "لا توجد حقول")
    await db.questions.update_one({"id": qid}, {"$set": upd})  # noqa: section stored as-is
    q = await db.questions.find_one({"id": qid}, {"_id": 0})
    if not q: raise HTTPException(404, "غير موجود")
    return q


@api.delete("/questions/{qid}")
async def delete_question(qid: str, admin=Depends(require_admin)):
    q = await db.questions.find_one({"id": qid}, {"_id": 0, "text": 1})
    res = await db.questions.delete_one({"id": qid})
    if res.deleted_count == 0: raise HTTPException(404, "غير موجود")
    await _audit(admin["id"], admin["name"], "delete", "question", qid,
                 (q.get("text","")[:80] if q else ""))
    return {"ok": True}


@api.post("/admin/seed-chiller-questions")
async def seed_chiller_questions(_=Depends(require_admin)):
    """One-time endpoint: replaces all chiller questions with the standard set."""
    CHILLER_QUESTIONS = [
        {"text": "Check temperature setting (must be 8–25°C)",          "answer_type": "numeric", "unit": "°C"},
        {"text": "Check actual temperature – High",                      "answer_type": "numeric", "unit": "°C"},
        {"text": "Check actual temperature – Low",                       "answer_type": "numeric", "unit": "°C"},
        {"text": "Check water pump pressure – High",                     "answer_type": "numeric", "unit": "kg"},
        {"text": "Check water pump pressure – Low",                      "answer_type": "numeric", "unit": "kg"},
        {"text": "Check any abnormal sound in the unit",                 "answer_type": "yes_no",  "unit": None},
        {"text": "Clean air filter if necessary",                        "answer_type": "yes_no",  "unit": None},
        {"text": "Check cooling fan if working (air cooled chiller)",    "answer_type": "yes_no",  "unit": None},
        {"text": "Check any water leakage",                              "answer_type": "yes_no",  "unit": None},
        {"text": "Check all water valves",                               "answer_type": "yes_no",  "unit": None},
        {"text": "Clean water quality in the tank, change if necessary", "answer_type": "yes_no",  "unit": None},
        {"text": "Fill up with water + Anti-rust Chemical",              "answer_type": "yes_no",  "unit": None},
    ]
    deleted = await db.questions.delete_many({"category": "chiller"})
    now = datetime.now(timezone.utc).isoformat()
    docs = [
        {"id": str(uuid.uuid4()), "category": "chiller", "text": q["text"],
         "order": i, "answer_type": q["answer_type"], "unit": q["unit"], "created_at": now}
        for i, q in enumerate(CHILLER_QUESTIONS)
    ]
    await db.questions.insert_many(docs)
    return {"deleted": deleted.deleted_count, "inserted": len(docs)}


@api.post("/admin/seed-mechanical-questions")
async def seed_mechanical_questions(_=Depends(require_admin)):
    """One-time endpoint: replaces all mechanical machine questions with the standard set."""
    MECHANICAL_QUESTIONS = [
        {"text": "Toggle bolts"},
        {"text": "Oil leakage"},
        {"text": "M/C saddle"},
        {"text": "Oil temperature"},
        {"text": "Tie bars"},
        {"text": "Hydraulic oil level"},
        {"text": "Ejector butterfly"},
        {"text": "Abnormal sound or movement"},
        {"text": "Hopper water"},
    ]
    deleted = await db.questions.delete_many({"category": "mechanical"})
    now = datetime.now(timezone.utc).isoformat()
    docs = [
        {"id": str(uuid.uuid4()), "category": "mechanical", "text": q["text"],
         "order": i, "answer_type": "yes_no", "unit": None, "created_at": now}
        for i, q in enumerate(MECHANICAL_QUESTIONS)
    ]
    await db.questions.insert_many(docs)
    return {"deleted": deleted.deleted_count, "inserted": len(docs)}


@api.post("/admin/seed-cooling-tower-questions")
async def seed_cooling_tower_questions(_=Depends(require_admin)):
    """One-time endpoint: replaces all cooling tower questions with the standard set."""
    CT_QUESTIONS = [
        {"text": "Check main tank water level"},
        {"text": "Check water distribution over cooling pad"},
        {"text": "Check exhaust fan operation"},
        {"text": "Check main tank water supply level"},
    ]
    deleted = await db.questions.delete_many({"category": "cooling_tower"})
    now = datetime.now(timezone.utc).isoformat()
    docs = [
        {"id": str(uuid.uuid4()), "category": "cooling_tower", "text": q["text"],
         "order": i, "answer_type": "yes_no", "unit": None, "created_at": now}
        for i, q in enumerate(CT_QUESTIONS)
    ]
    await db.questions.insert_many(docs)
    return {"deleted": deleted.deleted_count, "inserted": len(docs)}


@api.post("/admin/seed-preventive-questions")
async def seed_preventive_questions(_=Depends(require_admin)):
    """One-time endpoint: replaces all preventive maintenance questions with the 45-question set."""
    PREVENTIVE_QUESTIONS = [
        # Clamp Unit (18)
        {"section": "Clamp Unit", "text": "Check hydraulic oil level"},
        {"section": "Clamp Unit", "text": "Check hydraulic valves condition"},
        {"section": "Clamp Unit", "text": "Check pipes and hoses for leakage"},
        {"section": "Clamp Unit", "text": "Check hydraulic oil temperature"},
        {"section": "Clamp Unit", "text": "Check front/rear safety doors"},
        {"section": "Clamp Unit", "text": "Remove old grease from guide shoes"},
        {"section": "Clamp Unit", "text": "Grease nipples and rods of guide shoes"},
        {"section": "Clamp Unit", "text": "Check moving platen guide shoes"},
        {"section": "Clamp Unit", "text": "Check for water/oil leakage"},
        {"section": "Clamp Unit", "text": "Check tie bars and chains condition"},
        {"section": "Clamp Unit", "text": "Check heat exchanger"},
        {"section": "Clamp Unit", "text": "Check ejector piston and butterfly"},
        {"section": "Clamp Unit", "text": "Check door safety switches"},
        {"section": "Clamp Unit", "text": "Check for loose/broken bolts"},
        {"section": "Clamp Unit", "text": "Check wiring and cables"},
        {"section": "Clamp Unit", "text": "Check clamp/ejector scale"},
        {"section": "Clamp Unit", "text": "Lubricate all moving parts"},
        {"section": "Clamp Unit", "text": "Check for abnormal noise"},
        # Injection Unit (14)
        {"section": "Injection Unit", "text": "Check injection piston cylinders"},
        {"section": "Injection Unit", "text": "Check oil leakage in pipes and hoses"},
        {"section": "Injection Unit", "text": "Check manifold valves"},
        {"section": "Injection Unit", "text": "Check motor and pump condition"},
        {"section": "Injection Unit", "text": "Check carriage bolts and nuts"},
        {"section": "Injection Unit", "text": "Check barrel water circulation"},
        {"section": "Injection Unit", "text": "Check hydraulic motor"},
        {"section": "Injection Unit", "text": "Check nozzle heater"},
        {"section": "Injection Unit", "text": "Check barrel heaters and thermocouples"},
        {"section": "Injection Unit", "text": "Check lubrication nipples"},
        {"section": "Injection Unit", "text": "Lubricate motor and moving parts"},
        {"section": "Injection Unit", "text": "Check purge shield"},
        {"section": "Injection Unit", "text": "Check injection unit cables and wiring"},
        {"section": "Injection Unit", "text": "Check injection level and nozzle centralization"},
        # Electrical Panel (11)
        {"section": "Electrical Panel", "text": "Clean motor and wiring"},
        {"section": "Electrical Panel", "text": "Check emergency stop buttons"},
        {"section": "Electrical Panel", "text": "Check display and control wiring"},
        {"section": "Electrical Panel", "text": "Check cooling fans operation"},
        {"section": "Electrical Panel", "text": "Check voltage measurement"},
        {"section": "Electrical Panel", "text": "Check circuit breakers"},
        {"section": "Electrical Panel", "text": "Check contactors and relays"},
        {"section": "Electrical Panel", "text": "Check cables and wiring connections"},
        {"section": "Electrical Panel", "text": "Clean dust from electrical cabinet"},
        {"section": "Electrical Panel", "text": "Check battery and power supply"},
        {"section": "Electrical Panel", "text": "Check servo motor driver"},
        # General (2)
        {"section": "General", "text": "Check machine level"},
        {"section": "General", "text": "Clean machine exterior"},
    ]
    deleted = await db.questions.delete_many({"category": "preventive"})
    now = datetime.now(timezone.utc).isoformat()
    docs = [
        {"id": str(uuid.uuid4()), "category": "preventive", "section": q["section"],
         "text": q["text"], "order": i, "answer_type": "yes_no", "unit": None, "created_at": now}
        for i, q in enumerate(PREVENTIVE_QUESTIONS)
    ]
    await db.questions.insert_many(docs)
    return {"deleted": deleted.deleted_count, "inserted": len(docs)}


# ---------- Inspections ----------
COOLDOWN_MINUTES = 15


def _parse_iso(s: str) -> datetime:
    return datetime.fromisoformat(s.replace("Z", "+00:00"))


async def _get_last_inspection(target_id: str, category: str):
    return await db.inspections.find_one(
        {"target_id": target_id, "category": category},
        {"_id": 0},
        sort=[("created_at", -1)],
    )


@api.get("/inspections/cooldown")
async def cooldown(target_id: str, category: str, _=Depends(get_current_user)):
    last = await _get_last_inspection(target_id, category)
    if not last:
        return {"in_cooldown": False, "remaining_seconds": 0, "last_at": None}
    last_at = _parse_iso(last["created_at"])
    now = datetime.now(timezone.utc)
    elapsed = (now - last_at).total_seconds()
    cooldown_total = COOLDOWN_MINUTES * 60
    remaining = max(0, int(cooldown_total - elapsed))
    return {
        "in_cooldown": remaining > 0,
        "remaining_seconds": remaining,
        "cooldown_seconds": cooldown_total,
        "last_at": last["created_at"],
        "last_technician_name": last.get("technician_name", ""),
    }


LOCK_HOURS = 12

@api.get("/inspections/locked-targets")
async def locked_targets(category: str, _=Depends(get_current_user)):
    """Return target_ids inspected by any technician within the last 12 hours."""
    if category in NO_COOLDOWN_CATS:
        return {"locked": []}
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=LOCK_HOURS)).isoformat()
    docs = await db.inspections.find(
        {"category": category, "created_at": {"$gte": cutoff}},
        {"_id": 0, "target_id": 1},
    ).to_list(1000)
    return {"locked": list({d["target_id"] for d in docs})}


@api.post("/inspections")
async def create_inspection(body: InspectionCreate, user=Depends(get_current_user)):
    if body.category not in allowed_categories(user):
        raise HTTPException(403, "لا تملك صلاحية هذا القسم")
    expected = CAT_TARGET[body.category]
    if body.target_type != expected:
        raise HTTPException(400, "نوع الهدف لا يطابق القسم")
    coll = {"machine": "machines", "chiller": "chillers", "panel": "panels", "cooling_tower": "cooling_towers"}[body.target_type]
    target = await db[coll].find_one({"id": body.target_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "العنصر غير موجود")

    # 15-minute cooldown per (target, category) — skipped for NO_COOLDOWN_CATS
    if body.category not in NO_COOLDOWN_CATS:
        last = await _get_last_inspection(body.target_id, body.category)
        if last:
            elapsed = (datetime.now(timezone.utc) - _parse_iso(last["created_at"])).total_seconds()
            remaining = COOLDOWN_MINUTES * 60 - elapsed
            if remaining > 0:
                mins = int(remaining // 60) + 1
                raise HTTPException(
                    429,
                    f"لا يمكن رفع فحص جديد لنفس العنصر قبل مرور {COOLDOWN_MINUTES} دقيقة. تبقّى ~{mins} دقيقة.",
                )

    doc = {"id": str(uuid.uuid4()), "category": body.category,
           "target_type": body.target_type, "target_id": body.target_id,
           "target_number": target["number"], "target_name": target.get("name", ""),
           "technician_id": user["id"], "technician_name": user["name"],
           "answers": [
               {k: v for k, v in {
                   "qid": a.question_id, "a": a.answer,
                   **({"nv": a.numeric_value} if a.numeric_value is not None else {}),
                   **({"n": a.note} if a.note else {}),
                   **({"sp": a.spare_part} if a.spare_part else {}),
               }.items()}
               for a in body.answers
           ], "notes": body.notes or "",
           "time_mc_received": body.time_mc_received or "",
           "time_delivered": body.time_delivered or "",
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.inspections.insert_one(doc); doc.pop("_id", None)
    return doc


@api.get("/inspections")
async def list_inspections(
    target_number: Optional[str] = None, target_type: Optional[str] = None,
    category: Optional[str] = None, technician_id: Optional[str] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    limit: int = 200, user=Depends(get_current_user),
):
    q: dict = {}
    if user["role"] == "admin":
        if technician_id:
            q["technician_id"] = technician_id
    else:
        q["technician_id"] = user["id"]
        if user["role"] == "helper":
            q["category"] = "preventive"
    if target_number: q["target_number"] = {"$regex": f"^{re.escape(target_number)}", "$options": "i"}
    if target_type: q["target_type"] = target_type
    if category: q["category"] = category
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    docs = await db.inspections.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    for d in docs:
        if d.get("answers"):
            d["answers"] = _expand_answers(d["answers"])
    return docs


@api.get("/inspections/export/csv")
async def export_csv(request: Request, target_number: Optional[str] = None, target_type: Optional[str] = None,
                     category: Optional[str] = None, date_from: Optional[str] = None,
                     date_to: Optional[str] = None, _=Depends(require_admin)):
    _rl_check_export(request.client.host)
    q: dict = {}
    if target_number: q["target_number"] = {"$regex": f"^{re.escape(target_number)}", "$options": "i"}
    if target_type: q["target_type"] = target_type
    if category: q["category"] = category
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    items = await db.inspections.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    questions = await db.questions.find({}, {"_id": 0}).to_list(2000)
    qmap = {x["id"]: x for x in questions}
    buf = io.StringIO(); buf.write("\ufeff")
    buf.write("التاريخ,النوع,الرقم,الاسم,الفني,القسم,السؤال,الإجابة,ملاحظة\n")
    type_map = {"machine": "مكينة", "chiller": "شيلر", "panel": "لوحة"}
    cat_map = {"electrical": "كهرباء", "mechanical": "ميكانيكا", "chiller": "شيلر", "panels": "لوحات"}
    for insp in items:
        for ans in insp.get("answers", []):
            qobj = qmap.get(ans.get("qid") or ans.get("question_id"), {})
            text = (qobj.get("text", "") or "").replace(",", "،").replace("\n", " ")
            note = (ans.get("n") or ans.get("note") or "").replace(",", "،").replace("\n", " ")
            ans_val = ans["a"] if "a" in ans else ans.get("answer")
            row = (f'{insp["created_at"]},{type_map.get(insp.get("target_type",""),"")},'
                   f'{insp.get("target_number","")},{insp.get("target_name","")},'
                   f'{insp["technician_name"]},{cat_map.get(insp.get("category",""),"")},'
                   f'{text},{"صح" if ans_val else "خطأ"},{note}\n')
            buf.write(row)
    return StreamingResponse(iter([buf.getvalue().encode("utf-8")]),
                             media_type="text/csv; charset=utf-8",
                             headers={"Content-Disposition": 'attachment; filename="inspections.csv"'})


@api.get("/inspections/export/excel")
async def export_excel(
    request: Request,
    target_number: Optional[str] = None,
    target_type: Optional[str] = None,
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    admin_user=Depends(require_admin),
):
    _rl_check_export(request.client.host)
    from datetime import date as date_cls, timedelta
    import calendar

    if not target_number:
        raise HTTPException(400, "يجب تحديد رقم المعدة")
    if not date_from or not date_to:
        raise HTTPException(400, "يجب تحديد نطاق التاريخ")

    IS_PANEL = category in ("panels_main", "panels_sub")

    q: dict = {}
    if target_number: q["target_number"] = {"$regex": f"^{re.escape(target_number)}", "$options": "i"}
    if target_type: q["target_type"] = target_type
    if category: q["category"] = category
    q["created_at"] = {"$gte": date_from, "$lte": date_to + "T23:59:59"}

    items = await db.inspections.find(q, {"_id": 0}).sort("created_at", 1).to_list(5000)
    questions_list = await db.questions.find(
        {"category": category} if category else {}, {"_id": 0}
    ).sort([("category", 1), ("order", 1)]).to_list(2000)

    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")
    PURPLE       = "6B2D6B"
    PURPLE_LIGHT = "F3E8F3"
    WHITE        = "FFFFFF"
    GREEN_BG     = "E8F5E9"
    RED_BG       = "FFEBEE"
    GREEN_FG     = "2E7D32"
    RED_FG       = "C62828"
    HEADER_BG    = "4A1442"
    DAY_BG       = "EDE7F6"
    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color=PURPLE)
    def cb(top=thin, bottom=thin, left=thin, right=thin):
        return Border(top=top, bottom=bottom, left=left, right=right)

    d_from = date_cls.fromisoformat(date_from)
    d_to   = date_cls.fromisoformat(date_to)

    cat_label = {
        "electrical": "Electrical", "mechanical": "Mechanical",
        "chiller": "Chiller", "panels_main": "Main Panels", "panels_sub": "Sub Panels"
    }.get(category or "", "")
    type_label = {"machine": "Machine", "chiller": "Chiller", "panel": "Panel"}.get(target_type or "", "")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def build_sheet(ws, insp_by_key, columns, col_label_fn, total_cols):
        last_col = get_column_letter(total_cols)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 60
        for ci in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 15 if IS_PANEL else 7

        # Row 1: top spacing | Row 2: logo full-width | Row 3: checklist title
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 38
        for r in range(1, 5):
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=WHITE)
                cell.border = Border()
        ws.merge_cells(f"A2:{last_col}2")
        if os.path.exists(LOGO_PATH):
            img = XLImage(LOGO_PATH)
            other_col_px = int((15 if IS_PANEL else 7) * 7)
            img.width = int(60 * 7) + (total_cols - 1) * other_col_px
            img.height = 82
            img.anchor = "A2"
            ws.add_image(img)
        ws.merge_cells(f"A3:{last_col}3")
        freq = "Weekly" if IS_PANEL else "Daily"
        tc = ws["A3"]
        tc.value = f"{freq} {cat_label} Inspection Checklist"
        tc.font = Font(name="Arial", bold=True, size=16, color=PURPLE)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = Border(bottom=Side(style="medium", color=PURPLE))

        # Row 4: info bar
        ws.row_dimensions[4].height = 28
        ws.merge_cells(f"A4:{last_col}4")
        ic = ws["A4"]
        if IS_PANEL:
            ic.value = f"   Panel #: {target_number}     Section: {cat_label}     Period: {d_from.strftime('%B %Y')}"
        else:
            ic.value = (f"   {type_label} #: {target_number}     Section: {cat_label}     "
                        f"Period: {d_from.strftime('%d %b %Y')} to {d_to.strftime('%d %b %Y')}")
        ic.font = Font(name="Arial", bold=True, size=12, color=WHITE)
        ic.fill = PatternFill("solid", fgColor=HEADER_BG)
        ic.alignment = Alignment(horizontal="left", vertical="center")

        # Row 5: period header
        ws.row_dimensions[5].height = 22
        if IS_PANEL:
            ws.merge_cells(f"B5:{last_col}5")
            mc = ws["B5"]
            mc.value = d_from.strftime("%B %Y")
            mc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
            mc.fill = PatternFill("solid", fgColor="7B3F7B")
            mc.alignment = Alignment(horizontal="center", vertical="center")
            mc.border = cb(left=thick, right=thick)
        else:
            month_groups: dict = {}
            for ci, d in enumerate(columns, start=2):
                month_groups.setdefault((d.year, d.month), []).append(ci)
            for (yr, mo), cols in month_groups.items():
                s = get_column_letter(cols[0]); e = get_column_letter(cols[-1])
                ws.merge_cells(f"{s}5:{e}5")
                mc2 = ws[f"{s}5"]
                mc2.value = date_cls(yr, mo, 1).strftime("%B %Y")
                mc2.font = Font(name="Arial", bold=True, size=10, color=WHITE)
                mc2.fill = PatternFill("solid", fgColor="7B3F7B")
                mc2.alignment = Alignment(horizontal="center", vertical="center")
                mc2.border = cb(left=thick, right=thick)

        # Row 6: column headers
        ws.row_dimensions[6].height = 28
        qh = ws["A6"]
        qh.value = "Check Points"
        qh.font = Font(name="Arial", bold=True, size=12, color=WHITE)
        qh.fill = PatternFill("solid", fgColor=HEADER_BG)
        qh.alignment = Alignment(horizontal="center", vertical="center")
        qh.border = cb(left=thick, right=thick, top=thick, bottom=thick)
        for ci, col in enumerate(columns, start=2):
            cl = get_column_letter(ci)
            dc = ws[f"{cl}6"]
            dc.value = col_label_fn(col)
            dc.font = Font(name="Arial", bold=True, size=10, color=PURPLE)
            dc.fill = PatternFill("solid", fgColor=DAY_BG)
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.border = cb()

        # Questions
        for qi, q in enumerate(questions_list):
            row = 7 + qi
            ws.row_dimensions[row].height = 34
            q_is_numeric = q.get("answer_type") == "numeric"
            qc = ws[f"A{row}"]
            qc.value = q["text"]
            qc.font = Font(name="Arial", size=11)
            qc.fill = PatternFill("solid", fgColor=PURPLE_LIGHT if qi % 2 == 0 else WHITE)
            qc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            qc.border = cb(left=thick, right=thick)
            for ci, col in enumerate(columns, start=2):
                cl = get_column_letter(ci)
                cell = ws[f"{cl}{row}"]
                key = col if IS_PANEL else col.isoformat()
                insp = insp_by_key.get(key)
                cell_val = None
                if insp:
                    ans_map = {(a.get("qid") or a.get("question_id")): a for a in insp.get("answers", [])}
                    ans = ans_map.get(q["id"])
                    if ans:
                        if q_is_numeric:
                            nv = ans["nv"] if "nv" in ans else ans.get("numeric_value")
                            if nv is not None: cell_val = ("numeric", nv)
                        else:
                            raw = ans["a"] if "a" in ans else ans.get("answer")
                            if (ans.get("sk") or ans.get("skipped")) or raw is None:
                                cell_val = ("na", None)
                            else:
                                cell_val = ("bool", raw)
                if cell_val:
                    if cell_val[0] == "numeric":
                        cell.value = f"{cell_val[1]}°C"
                        cell.font = Font(name="Arial", bold=True, size=11, color=PURPLE)
                        cell.fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
                    elif cell_val[0] == "na":
                        cell.value = "N/A"
                        cell.font = Font(name="Arial", bold=True, size=10, color="64748B")
                        cell.fill = PatternFill("solid", fgColor="E2E8F0")
                    else:
                        ok = cell_val[1]
                        cell.value = "✓" if ok else "✗"
                        cell.font = Font(name="Arial", bold=True, size=14,
                                         color=GREEN_FG if ok else RED_FG)
                        cell.fill = PatternFill("solid", fgColor=GREEN_BG if ok else RED_BG)
                else:
                    cell.fill = PatternFill("solid", fgColor=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = cb()

        # Footer
        last_q_row = 6 + len(questions_list)
        fr = last_q_row + 2
        ws.row_dimensions[fr].height = 24
        mid_ci = max(2, total_cols // 2)
        ws.merge_cells(f"A{fr}:{get_column_letter(mid_ci - 1)}{fr}")
        rb = ws[f"A{fr}"]
        rb.value = "Done by:"
        rb.font = Font(name="Arial", bold=True, size=11, color=PURPLE)
        rb.fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
        rb.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        rb.border = cb(left=thick, bottom=thick)
        ws.merge_cells(f"{get_column_letter(mid_ci)}{fr}:{last_col}{fr}")
        me = ws[f"{get_column_letter(mid_ci)}{fr}"]
        me.value = f"Maintenance Engineer:   {admin_user.get('name', '')}"
        me.font = Font(name="Arial", bold=True, size=11, color=PURPLE)
        me.fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
        me.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        me.border = cb(right=thick, bottom=thick)
        tr = fr + 1
        ws.row_dimensions[tr].height = 28
        tc2 = ws[f"A{tr}"]
        tc2.value = "Technician"
        tc2.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        tc2.fill = PatternFill("solid", fgColor=HEADER_BG)
        tc2.alignment = Alignment(horizontal="center", vertical="center")
        tc2.border = cb(left=thick, right=thick, bottom=thick)
        for ci, col in enumerate(columns, start=2):
            cl = get_column_letter(ci)
            cell = ws[f"{cl}{tr}"]
            key = col if IS_PANEL else col.isoformat()
            insp = insp_by_key.get(key)
            if insp:
                name = insp.get("technician_name", "")
                initials = "".join(p[0].upper() for p in name.split() if p)
                cell.value = initials
                cell.font = Font(name="Arial", size=10, bold=True, color=PURPLE)
            cell.fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cb(bottom=thick)

        ws.freeze_panes = "B7"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1

    # ── Build the sheet ──────────────────────────────────────────────────────
    if IS_PANEL:
        insp_by_week: dict = {}
        for insp in items:
            d = date_cls.fromisoformat(insp["created_at"][:10])
            first_day = date_cls(d.year, d.month, 1)
            week_num = (d.day + first_day.weekday()) // 7 + 1
            key = f"week_{week_num}"
            insp_by_week[key] = insp
        first_day = date_cls(d_from.year, d_from.month, 1)
        last_day_num = calendar.monthrange(d_from.year, d_from.month)[1]
        num_weeks = (last_day_num + first_day.weekday()) // 7 + 1
        weeks = [f"week_{i}" for i in range(1, num_weeks + 1)]
        week_labels = {f"week_{i}": f"Week {i}" for i in range(1, num_weeks + 1)}
        ws = wb.create_sheet(title=f"{target_number}"[:31])
        build_sheet(ws, insp_by_week, weeks, lambda w: week_labels[w], len(weeks) + 1)
    else:
        days = [d_from + timedelta(i) for i in range((d_to - d_from).days + 1)]
        insp_by_date: dict = {}
        for insp in items:
            insp_by_date[insp["created_at"][:10]] = insp
        ws = wb.create_sheet(title=f"{target_number}"[:31])
        build_sheet(ws, insp_by_date, days, lambda d: d.day, len(days) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"inspection_{target_number}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api.get("/inspections/{iid}")
async def get_inspection(iid: str, user=Depends(get_current_user)):
    doc = await db.inspections.find_one({"id": iid}, {"_id": 0})
    if not doc: raise HTTPException(404, "غير موجود")
    if user["role"] != "admin" and doc["technician_id"] != user["id"]:
        raise HTTPException(403, "غير مصرح")
    if doc.get("answers"):
        doc["answers"] = _expand_answers(doc["answers"])
    return doc


@api.patch("/inspections/{iid}")
async def update_inspection(iid: str, body: InspectionUpdate, user=Depends(get_current_user)):
    doc = await db.inspections.find_one({"id": iid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "غير موجود")
    if doc["technician_id"] != user["id"]:
        raise HTTPException(403, "يمكنك تعديل فحوصاتك فقط")
    elapsed = (datetime.now(timezone.utc) - _parse_iso(doc["created_at"])).total_seconds()
    if elapsed > 3600:
        raise HTTPException(403, "انتهت مدة التعديل المسموح بها — ساعة واحدة من وقت الرفع")
    new_answers = [
        {k: v for k, v in {
            "qid": a.question_id, "a": a.answer,
            **({"nv": a.numeric_value} if a.numeric_value is not None else {}),
            **({"n": a.note} if a.note else {}),
            **({"sp": a.spare_part} if a.spare_part else {}),
        }.items()}
        for a in body.answers
    ]
    await db.inspections.update_one(
        {"id": iid},
        {"$set": {"answers": new_answers, "notes": body.notes or "",
                  "time_mc_received": body.time_mc_received or "",
                  "time_delivered": body.time_delivered or "",
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    updated = await db.inspections.find_one({"id": iid}, {"_id": 0})
    if updated.get("answers"):
        updated["answers"] = _expand_answers(updated["answers"])
    return updated


@api.delete("/inspections/{iid}")
async def delete_inspection(iid: str, _=Depends(require_admin)):
    res = await db.inspections.delete_one({"id": iid})
    if res.deleted_count == 0: raise HTTPException(404, "غير موجود")
    return {"ok": True}


@api.get("/inspections/{iid}/export/pdf")
async def export_pdf(iid: str, _=Depends(require_admin)):
    doc = await db.inspections.find_one({"id": iid}, {"_id": 0})
    if not doc: raise HTTPException(404, "غير موجود")
    questions = await db.questions.find({}, {"_id": 0}).to_list(2000)
    qmap = {x["id"]: x for x in questions}
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 2 * cm
    type_label = {"machine": "Machine", "chiller": "Chiller", "panel": "Panel"}.get(doc.get("target_type", ""), "")
    cat_label = {"electrical": "Electrical", "mechanical": "Mechanical", "chiller": "Chiller", "panels": "Panels"}.get(doc.get("category", ""), "")
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2 * cm, y, f"Inspection Report - {type_label} #{doc.get('target_number','')}")
    y -= 0.7 * cm; c.setFont("Helvetica", 10)
    c.drawString(2 * cm, y, f"Section: {cat_label}"); y -= 0.5 * cm
    c.drawString(2 * cm, y, f"Date: {doc['created_at']}"); y -= 0.5 * cm
    c.drawString(2 * cm, y, f"Technician: {doc['technician_name']}"); y -= 0.8 * cm
    for a in doc.get("answers", []):
        q = qmap.get(a.get("qid") or a.get("question_id"), {})
        mark = "OK" if (a["a"] if "a" in a else a.get("answer")) else "FAIL"
        c.drawString(2.4 * cm, y, f"[{mark}] {q.get('text','')[:80]}")
        y -= 0.5 * cm
        if y < 2 * cm:
            c.showPage(); y = height - 2 * cm
    c.showPage(); c.save(); buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="inspection_{doc.get("target_number","")}.pdf"'})


@api.get("/inspections/{iid}/export/preventive-pdf")
async def export_preventive_pdf(iid: str, _=Depends(require_admin)):
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        Image as PLImage,
    )
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    doc_data = await db.inspections.find_one({"id": iid}, {"_id": 0})
    if not doc_data: raise HTTPException(404, "غير موجود")
    if doc_data.get("category") != "preventive":
        raise HTTPException(400, "هذا ليس فحص صيانة وقائية")

    questions = await db.questions.find(
        {"category": "preventive"}, {"_id": 0}
    ).sort([("order", 1)]).to_list(200)

    expanded = _expand_answers(doc_data.get("answers", []))
    answers_map = {a["question_id"]: a for a in expanded}

    sections: dict = {}
    section_order: list = []
    for q in questions:
        sec = q.get("section") or "General"
        if sec not in sections:
            sections[sec] = []
            section_order.append(sec)
        sections[sec].append(q)

    # ── Colors ────────────────────────────────────────────────────
    PURPLE       = rl_colors.HexColor('#6B2D6B')
    DARK_PURPLE  = rl_colors.HexColor('#4A1442')
    PURPLE_LIGHT = rl_colors.HexColor('#EDE0ED')
    WHITE        = rl_colors.white
    BLACK        = rl_colors.black
    GREEN_FG     = rl_colors.HexColor('#1B5E20')
    GREEN_BG     = rl_colors.HexColor('#E8F5E9')
    RED_FG       = rl_colors.HexColor('#B71C1C')
    RED_BG       = rl_colors.HexColor('#FFEBEE')
    GREY         = rl_colors.HexColor('#757575')
    LIGHT_GREY   = rl_colors.HexColor('#F5F5F5')
    MID_GREY     = rl_colors.HexColor('#CCCCCC')
    SP_BG        = rl_colors.HexColor('#FFF8E1')

    date_str          = doc_data.get("created_at", "")[:10]
    machine_no        = doc_data.get("target_number", "")
    machine_name      = doc_data.get("target_name", "") or ""
    tech_name         = doc_data.get("technician_name", "")
    notes_text        = doc_data.get("notes", "") or ""
    time_received     = doc_data.get("time_mc_received", "") or "—"
    time_delivered    = doc_data.get("time_delivered", "") or "—"

    buf = io.BytesIO()
    pw, ph = A4
    LM = RM = 1.4 * cm
    USABLE = pw - LM - RM   # ~18.2 cm

    # ── Paragraph helpers ─────────────────────────────────────────
    def ps(name, font="Helvetica", size=9, bold=False, align=TA_LEFT, color=BLACK, leading=None):
        return ParagraphStyle(
            name, fontName="Helvetica-Bold" if bold else font,
            fontSize=size, leading=leading or size * 1.35,
            textColor=color, alignment=align,
        )

    p_title   = ps("t",  size=13, bold=True,  color=PURPLE,    align=TA_CENTER)
    p_sub     = ps("s",  size=10, bold=True,  color=DARK_PURPLE, align=TA_CENTER)
    p_info    = ps("i",  size=9)
    p_hdr     = ps("h",  size=8,  bold=True,  color=WHITE,     align=TA_CENTER)
    p_sec_hdr = ps("sh", size=9,  bold=True,  color=WHITE)
    p_num     = ps("n",  size=8,  align=TA_CENTER)
    p_text    = ps("tx", size=8)
    p_ok      = ps("ok", size=12, bold=True,  color=GREEN_FG,  align=TA_CENTER)
    p_fail    = ps("fl", size=12, bold=True,  color=RED_FG,    align=TA_CENTER)
    p_na      = ps("na", size=7,  bold=True,  color=GREY,      align=TA_CENTER)
    p_sp      = ps("sp", size=7,  color=GREY)
    p_label   = ps("lb", size=9,  bold=True,  color=DARK_PURPLE)

    # ── Document ──────────────────────────────────────────────────
    doc_obj = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=LM, rightMargin=RM,
        topMargin=1.2 * cm, bottomMargin=1.5 * cm,
    )
    story = []

    # ── HEADER ────────────────────────────────────────────────────
    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")

    story_items = []

    # Logo — full width, centered, tall
    if os.path.exists(LOGO_PATH):
        logo = PLImage(LOGO_PATH, width=USABLE, height=3.2*cm, kind='proportional')
        story_items.append(logo)
        story_items.append(Spacer(1, 0.1*cm))

    # Title block
    title_tbl = Table(
        [
            [Paragraph("Hydraulic Injection Machine", p_title)],
            [Paragraph("PREVENTIVE MAINTENANCE CHECKLIST", p_sub)],
        ],
        colWidths=[USABLE],
    )
    title_tbl.setStyle(TableStyle([
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND',    (0,0), (-1,-1), LIGHT_GREY),
        ('BOX',           (0,0), (-1,-1), 1.5, PURPLE),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story_items.append(title_tbl)
    story_items.append(Spacer(1, 0.2*cm))

    # Info row
    info_tbl = Table(
        [
            [
                Paragraph(f"<b>Date:</b>  {date_str}", p_info),
                Paragraph(f"<b>Machine No.:</b>  {machine_no}" + (f"  &nbsp; <b>Model:</b>  {machine_name}" if machine_name else ""), p_info),
            ],
            [
                Paragraph("<b>Shift:</b>  —", p_info),
                Paragraph(f"<b>Maintenance Engineer:</b>  {tech_name}", p_info),
            ],
            [
                Paragraph(f"<b>Time Received:</b>  {time_received}", p_info),
                Paragraph(f"<b>Time Delivered:</b>  {time_delivered}", p_info),
            ],
        ],
        colWidths=[USABLE * 0.4, USABLE * 0.6],
    )
    info_tbl.setStyle(TableStyle([
        ('BOX',           (0,0), (-1,-1), 1.2, PURPLE),
        ('INNERGRID',     (0,0), (-1,-1), 0.3, MID_GREY),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
    ]))
    story_items.append(info_tbl)
    story_items.append(Spacer(1, 0.3*cm))

    for item in story_items:
        story.append(item)

    # ── CHECKLIST TABLE  (4 columns) ──────────────────────────────
    # Columns: # | CHECK POINTS | SPARE PART | OK/FAIL
    COL_N  = 0.70 * cm
    COL_R  = 1.90 * cm
    COL_SP = 3.40 * cm
    COL_T  = USABLE - COL_N - COL_R - COL_SP
    col_ws = [COL_N, COL_T, COL_SP, COL_R]

    rows   = []
    tstyle = [
        ('BOX',          (0, 0), (-1, -1), 1.2, PURPLE),
        ('INNERGRID',    (0, 0), (-1, -1), 0.3, MID_GREY),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',   (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ('LEFTPADDING',  (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]

    # Column header row
    rows.append([
        Paragraph("م", p_hdr),
        Paragraph("CHECK  POINTS", p_hdr),
        Paragraph("SPARE  PART", p_hdr),
        Paragraph("OK / FAIL", p_hdr),
    ])
    tstyle += [
        ('BACKGROUND', (0, 0), (-1, 0), DARK_PURPLE),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('LINEBELOW',  (0, 0), (-1, 0), 1.5, PURPLE),
        ('ROWBACKGROUND', (0, 0), (-1, 0), DARK_PURPLE),
    ]

    row_i   = 1
    local_q = 0

    for sec in section_order:
        # Section header row — spans all 4 columns
        rows.append([
            Paragraph("", p_sec_hdr),
            Paragraph(f"  {sec.upper()}", p_sec_hdr),
            Paragraph("", p_sec_hdr),
            Paragraph("", p_sec_hdr),
        ])
        r = row_i
        tstyle += [
            ('SPAN',          (0, r), (-1, r)),
            ('BACKGROUND',    (0, r), (-1, r), PURPLE),
            ('LINEABOVE',     (0, r), (-1, r), 0.8, DARK_PURPLE),
            ('TOPPADDING',    (0, r), (-1, r), 5),
            ('BOTTOMPADDING', (0, r), (-1, r), 5),
        ]
        row_i += 1

        for i, q in enumerate(sections[sec]):
            local_q += 1
            a = answers_map.get(q["id"])

            # Result cell
            if a is None or a.get("skipped"):
                result_p = Paragraph("N/A", p_na)
                bg = None
            elif a.get("answer"):
                result_p = Paragraph("✓", p_ok)
                bg = GREEN_BG
            else:
                result_p = Paragraph("✗", p_fail)
                bg = RED_BG

            # Question text (with note sub-line)
            q_text = q.get("text", "")
            if a and a.get("note") and not a.get("skipped"):
                text_p = Paragraph(
                    f'{q_text}<br/><font size="7" color="#757575"> ↳ {a["note"]}</font>',
                    p_text,
                )
            else:
                text_p = Paragraph(q_text, p_text)

            # Spare part cell
            sp_val = (a.get("spare_part", "") if a else "") or ""
            sp_p   = Paragraph(sp_val, p_sp) if sp_val else Paragraph("", p_sp)

            rows.append([Paragraph(str(i + 1), p_num), text_p, sp_p, result_p])
            r = row_i
            if bg:
                tstyle.append(('BACKGROUND', (0, r), (-1, r), bg))
            elif sp_val:
                tstyle.append(('BACKGROUND', (2, r), (2, r), SP_BG))
            elif local_q % 2 == 0:
                tstyle.append(('BACKGROUND', (0, r), (-1, r), LIGHT_GREY))
            tstyle += [
                ('ALIGN', (0, r), (0, r), 'CENTER'),
                ('ALIGN', (3, r), (3, r), 'CENTER'),
            ]
            row_i += 1

    checklist_tbl = Table(rows, colWidths=col_ws, repeatRows=1)
    checklist_tbl.setStyle(TableStyle(tstyle))
    story.append(checklist_tbl)
    story.append(Spacer(1, 0.3 * cm))

    # ── REMARKS / SPARE PARTS SUMMARY ─────────────────────────────
    # Collect all spare parts used across all questions
    all_parts = [
        f"• [{a.get('question_id','')}] {a['spare_part']}"
        for a in expanded if a.get("spare_part")
    ]
    # Build question-id → text map for display
    qtext_map = {q["id"]: q.get("text", "") for q in questions}
    parts_lines = [
        f"• {qtext_map.get(a['question_id'], '')} → {a['spare_part']}"
        for a in expanded if a.get("spare_part")
    ]
    parts_str = "<br/>".join(parts_lines) if parts_lines else "—"

    rw = [3.2 * cm, USABLE - 3.2 * cm]
    remarks_tbl = Table(
        [
            [Paragraph("<b>Remarks:</b>", p_info), Paragraph(notes_text or "—", p_info)],
            [Paragraph("<b>Spare Parts Used:</b>", p_info),
             Paragraph(parts_str, ps("pt", size=8, color=GREY))],
        ],
        colWidths=rw,
    )
    remarks_tbl.setStyle(TableStyle([
        ('BOX',       (0, 0), (-1, -1), 0.8, PURPLE),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, MID_GREY),
        ('BACKGROUND',(0, 0), (0, -1),  PURPLE_LIGHT),
        ('VALIGN',    (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    story.append(remarks_tbl)
    story.append(Spacer(1, 0.3 * cm))

    # ── SIGNATURE / TIME ROW ──────────────────────────────────────
    sw = USABLE / 3.0
    sig_tbl = Table([[
        Paragraph(
            f"<b>TIME M/C Received:</b><br/><br/>"
            f"<font size='11'>{time_received}</font>",
            p_info,
        ),
        Paragraph(
            f"<b>Delivery Time:</b><br/><br/>"
            f"<font size='11'>{time_delivered}</font>",
            p_info,
        ),
        Paragraph(
            f"<b>Maintenance Engineer:</b><br/><br/>"
            f"<font size='10'>{tech_name}</font>",
            p_info,
        ),
    ]], colWidths=[sw, sw, sw])
    sig_tbl.setStyle(TableStyle([
        ('BOX',       (0, 0), (-1, -1), 0.8, PURPLE),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, MID_GREY),
        ('VALIGN',    (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND',(0, 0), (-1, 0),  LIGHT_GREY),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
    ]))
    story.append(sig_tbl)

    doc_obj.build(story)
    buf.seek(0)
    fname = f"preventive_{machine_no}_{date_str}.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _expand_answers(answers: list) -> list:
    result = []
    for a in answers:
        if "qid" in a:
            result.append({
                "question_id": a["qid"],
                "answer": a.get("a"),
                "numeric_value": a.get("nv"),
                "note": a.get("n", ""),
                "spare_part": a.get("sp", ""),
                "skipped": a.get("a") is None,
            })
        else:
            result.append(a)
    return result


def _open_failures_pipeline(
    target_number: Optional[str] = None,
    target_type: Optional[str] = None,
    category: Optional[str] = None,
    target_id: Optional[str] = None,
):
    """Latest answer per (target_id, question_id). Only keep those still failing."""
    pre_match: dict = {}
    if target_type:
        pre_match["target_type"] = target_type
    if category:
        pre_match["category"] = category
    if target_number:
        pre_match["target_number"] = {"$regex": f"^{re.escape(target_number)}", "$options": "i"}
    if target_id:
        pre_match["target_id"] = target_id

    pipe = []
    if pre_match:
        pipe.append({"$match": pre_match})
    pipe += [
        {"$unwind": "$answers"},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": {"target_id": "$target_id", "question_id": "$answers.qid"},
            "latest_answer": {"$first": "$answers.a"},
            "latest_note": {"$first": "$answers.n"},
            "latest_at": {"$first": "$created_at"},
            "inspection_id": {"$first": "$id"},
            "target_number": {"$first": "$target_number"},
            "target_type": {"$first": "$target_type"},
            "target_name": {"$first": "$target_name"},
            "category": {"$first": "$category"},
            "technician_name": {"$first": "$technician_name"},
        }},
        {"$match": {"latest_answer": False}},
        {"$sort": {"latest_at": -1}},
    ]
    return pipe


async def _count_open_failures() -> int:
    pipe = _open_failures_pipeline()
    pipe.append({"$count": "n"})
    async for d in db.inspections.aggregate(pipe):
        return d["n"]
    return 0


@api.get("/failures/open")
async def list_open_failures(
    target_number: Optional[str] = None,
    target_type: Optional[str] = None,
    category: Optional[str] = None,
    target_id: Optional[str] = None,
    user=Depends(get_current_user),
):
    if user["role"] not in ("admin", "helper"):
        raise HTTPException(403, "غير مصرح")
    if user["role"] == "helper":
        target_type = "machine"
    pipe = _open_failures_pipeline(target_number, target_type, category, target_id)
    questions = await db.questions.find({}, {"_id": 0, "id": 1, "text": 1}).to_list(2000)
    qmap = {q["id"]: q for q in questions}
    results = []
    async for d in db.inspections.aggregate(pipe):
        q = qmap.get(d["_id"]["question_id"], {})
        results.append({
            "target_id": d["_id"]["target_id"],
            "question_id": d["_id"]["question_id"],
            "question_text": q.get("text", ""),
            "target_number": d.get("target_number", ""),
            "target_name": d.get("target_name", ""),
            "target_type": d.get("target_type", ""),
            "category": d.get("category", ""),
            "technician_name": d.get("technician_name", ""),
            "inspection_id": d.get("inspection_id", ""),
            "note": d.get("latest_note", ""),
            "since": d.get("latest_at", ""),
        })
    return results


# ---------- Preventive Plans ----------
@api.post("/preventive-plans")
async def create_preventive_plan(body: PreventivePlanCreate, admin=Depends(require_admin)):
    machine = await db.machines.find_one({"id": body.machine_id}, {"_id": 0})
    if not machine:
        raise HTTPException(404, "المكينة غير موجودة")
    existing = await db.preventive_plans.find_one({
        "machine_id": body.machine_id, "scheduled_date": body.scheduled_date
    })
    if existing:
        raise HTTPException(400, "خطة موجودة مسبقاً لهذه المكينة في هذا اليوم")
    doc = {
        "id": str(uuid.uuid4()),
        "machine_id": body.machine_id,
        "machine_number": machine["number"],
        "machine_name": machine.get("name", ""),
        "scheduled_date": body.scheduled_date,
        "created_by": admin["id"],
        "created_by_name": admin["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.preventive_plans.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/preventive-plans")
async def list_preventive_plans(date: Optional[str] = None, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "helper"):
        raise HTTPException(403, "غير مصرح")
    q: dict = {}
    if user["role"] == "helper":
        today = datetime.now(timezone.utc).date().isoformat()
        q["scheduled_date"] = today
    elif date:
        q["scheduled_date"] = date
    docs = await db.preventive_plans.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs


@api.delete("/preventive-plans/{pid}")
async def delete_preventive_plan(pid: str, _=Depends(require_admin)):
    res = await db.preventive_plans.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(404, "الخطة غير موجودة")
    return {"ok": True}


def _calc_duration(start_str: str, end_str: str) -> str:
    """Parse 12h time strings (e.g. '10:30 PM') and return duration like '4h 30m'."""
    import re as _re
    def _p(s):
        if not s: return None
        m = _re.match(r'^(\d{1,2}):(\d{2})\s*(AM|PM)$', s.strip(), _re.IGNORECASE)
        if not m: return None
        h, mn, ap = int(m.group(1)), int(m.group(2)), m.group(3).upper()
        if ap == 'PM' and h != 12: h += 12
        if ap == 'AM' and h == 12: h = 0
        return h * 60 + mn
    s, e = _p(start_str), _p(end_str)
    if s is None or e is None: return "—"
    diff = e - s
    if diff < 0: diff += 1440  # cross-midnight (max 24h shift)
    return f"{diff // 60}h {diff % 60:02d}m"


# ---------- Breakdowns ----------
@api.post("/breakdowns")
async def create_breakdown(body: BreakdownCreate, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "technician"):
        raise HTTPException(403, "غير مصرح")
    machine = await db.machines.find_one({"id": body.machine_id}, {"_id": 0})
    if not machine:
        raise HTTPException(404, "المكينة غير موجودة")
    doc = {
        "id": str(uuid.uuid4()),
        "machine_id": body.machine_id,
        "machine_number": machine["number"],
        "machine_name": machine.get("name", ""),
        "brief_description": body.brief_description,
        "repair_description": body.repair_description or "",
        "start_time": body.start_time or "",
        "end_time": body.end_time or "",
        "technician_id": user["id"],
        "technician_name": user["name"],
        "status": "submitted",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.breakdowns.insert_one(doc)
    doc.pop("_id", None)
    await _audit(user["id"], user["name"], "create", "breakdown", doc["id"],
                 f"مكينة {machine['number']} — {body.brief_description[:60]}")
    await _notify("new_breakdown",
                  f"عطل جديد — مكينة {machine['number']}",
                  f"سجّل {user['name']} عطلاً: {body.brief_description[:80]}",
                  doc["id"])
    return doc


@api.get("/breakdowns")
async def list_breakdowns(
    machine_number: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    technician_id: Optional[str] = None,
    user=Depends(get_current_user),
):
    if user["role"] not in ("admin", "technician"):
        raise HTTPException(403, "غير مصرح")
    q: dict = {}
    if user["role"] == "technician":
        q["technician_id"] = user["id"]
    else:
        if technician_id:
            q["technician_id"] = technician_id
    if machine_number:
        q["machine_number"] = {"$regex": f"^{re.escape(machine_number)}", "$options": "i"}
    if status:
        q["status"] = status
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    docs = await db.breakdowns.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return docs


@api.patch("/breakdowns/{bid}/resolve")
async def resolve_breakdown(bid: str, admin=Depends(require_admin)):
    doc = await db.breakdowns.find_one({"id": bid})
    if not doc:
        raise HTTPException(404, "غير موجود")
    await db.breakdowns.update_one(
        {"id": bid},
        {"$set": {"status": "resolved", "resolved_at": datetime.now(timezone.utc).isoformat()}},
    )
    await _audit(admin["id"], admin["name"], "resolve", "breakdown", bid,
                 f"مكينة {doc.get('machine_number','')} — {doc.get('brief_description','')[:60]}")
    await _notify("breakdown_resolved",
                  f"تم حل العطل — مكينة {doc.get('machine_number','')}",
                  f"أغلق {admin['name']} عطل: {doc.get('brief_description','')[:80]}",
                  bid)
    return {"ok": True}


@api.delete("/breakdowns/{bid}")
async def delete_breakdown(bid: str, _=Depends(require_admin)):
    res = await db.breakdowns.delete_one({"id": bid})
    if res.deleted_count == 0:
        raise HTTPException(404, "غير موجود")
    return {"ok": True}


# ---------- Downtime Reasons ----------
@api.get("/downtime-reasons")
async def list_downtime_reasons(_=Depends(get_current_user)):
    docs = await db.downtime_reasons.find({}, {"_id": 0}).sort("order", 1).to_list(200)
    return docs


@api.post("/downtime-reasons")
async def create_downtime_reason(body: DowntimeReasonCreate, _=Depends(require_admin)):
    count = await db.downtime_reasons.count_documents({})
    doc = {
        "id": str(uuid.uuid4()),
        "text": body.text.strip(),
        "specialty": body.specialty,
        "order": count,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.downtime_reasons.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.patch("/downtime-reasons/{rid}")
async def update_downtime_reason(rid: str, body: DowntimeReasonUpdate, _=Depends(require_admin)):
    update: dict = {}
    if body.text is not None:
        update["text"] = body.text.strip()
    if body.specialty is not None:
        update["specialty"] = body.specialty
    if not update:
        raise HTTPException(400, "No fields to update")
    await db.downtime_reasons.update_one({"id": rid}, {"$set": update})
    doc = await db.downtime_reasons.find_one({"id": rid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    return doc


@api.delete("/downtime-reasons/{rid}")
async def delete_downtime_reason(rid: str, _=Depends(require_admin)):
    res = await db.downtime_reasons.delete_one({"id": rid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"ok": True}


@api.get("/breakdowns/export/excel")
async def export_breakdowns_excel(
    request: Request,
    machine_number: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    _=Depends(require_admin),
):
    _rl_check_export(request.client.host)
    q: dict = {}
    if machine_number:
        q["machine_number"] = {"$regex": f"^{re.escape(machine_number)}", "$options": "i"}
    if status:
        q["status"] = status
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    docs = await db.breakdowns.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)

    PURPLE       = "6B2D6B"
    PURPLE_LIGHT = "F3E8F3"
    WHITE        = "FFFFFF"
    HEADER_BG    = "4A1442"
    GREEN_BG     = "E8F5E9"
    GREEN_FG     = "2E7D32"
    AMBER_BG     = "FFF8E1"
    AMBER_FG     = "F57F17"
    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color=PURPLE)

    def cb(top=None, bottom=None, left=None, right=None):
        return Border(
            top=top    or thin,
            bottom=bottom or thin,
            left=left  or thin,
            right=right or thin,
        )

    HEADERS    = ["#", "Date", "Machine Number", "Downtime Reason", "Technician",
                  "Start Time", "End Time", "Duration", "Repair Description", "Status"]
    COL_WIDTHS = [5,   20,     16,               34,                22,
                  13,         13,        13,         40,                  13]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Machine Downtime"
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_col = get_column_letter(len(HEADERS))

    # Row 1 spacing
    ws.row_dimensions[1].height = 8
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=WHITE)

    # Row 2 logo
    ws.row_dimensions[2].height = 80
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].fill = PatternFill("solid", fgColor=WHITE)
    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width  = sum(COL_WIDTHS) * 7
        img.height = 82
        img.anchor = "A2"
        ws.add_image(img)

    # Row 3 title
    ws.row_dimensions[3].height = 38
    ws.merge_cells(f"A3:{last_col}3")
    tc = ws["A3"]
    tc.value = "Machine Downtime Report"
    tc.font  = Font(name="Arial", bold=True, size=16, color=PURPLE)
    tc.fill  = PatternFill("solid", fgColor=WHITE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = Border(bottom=Side(style="medium", color=PURPLE))

    # Row 4 info bar
    ws.row_dimensions[4].height = 28
    ws.merge_cells(f"A4:{last_col}4")
    period = ""
    if date_from and date_to:
        period = f"   |   Period: {date_from}  to  {date_to}"
    elif date_from:
        period = f"   |   From: {date_from}"
    ic = ws["A4"]
    ic.value = f"   Total Records: {len(docs)}{period}"
    ic.font  = Font(name="Arial", bold=True, size=12, color=WHITE)
    ic.fill  = PatternFill("solid", fgColor=HEADER_BG)
    ic.alignment = Alignment(horizontal="left", vertical="center")

    # Row 5 column headers
    ws.row_dimensions[5].height = 30
    for ci, hdr in enumerate(HEADERS, 1):
        cl   = get_column_letter(ci)
        cell = ws[f"{cl}5"]
        cell.value = hdr
        cell.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=thick, bottom=thick,
            left=thick if ci == 1 else thin,
            right=thick if ci == len(HEADERS) else thin,
        )

    # Data rows
    for ri, doc in enumerate(docs):
        row = 6 + ri
        ws.row_dimensions[row].height = 22
        bg    = PURPLE_LIGHT if ri % 2 == 0 else WHITE
        stat  = doc.get("status", "")
        dur   = _calc_duration(doc.get("start_time", ""), doc.get("end_time", ""))
        mname = doc.get("machine_number", "")
        if doc.get("machine_name"):
            mname += f" — {doc['machine_name']}"
        values = [
            ri + 1,
            doc.get("created_at", "")[:10],
            mname,
            doc.get("brief_description", ""),
            doc.get("technician_name", ""),
            doc.get("start_time", "") or "—",
            doc.get("end_time", "")   or "—",
            dur,
            doc.get("repair_description", "") or "—",
            "Resolved" if stat == "resolved" else "Open",
        ]
        for ci, val in enumerate(values, 1):
            cl   = get_column_letter(ci)
            cell = ws[f"{cl}{row}"]
            cell.value = val
            cell.font  = Font(name="Arial", size=10)
            cell.alignment = Alignment(
                horizontal="center" if ci in (1, 6, 7, 8, 10) else "left",
                vertical="center", wrap_text=(ci in (4, 9)),
            )
            cell.border = Border(
                top=thin, bottom=thin,
                left=thick if ci == 1 else thin,
                right=thick if ci == len(HEADERS) else thin,
            )
            if ci == 10:
                if stat == "resolved":
                    cell.font = Font(name="Arial", bold=True, size=10, color=GREEN_FG)
                    cell.fill = PatternFill("solid", fgColor=GREEN_BG)
                else:
                    cell.font = Font(name="Arial", bold=True, size=10, color=AMBER_FG)
                    cell.fill = PatternFill("solid", fgColor=AMBER_BG)
            elif ci == 8:
                cell.font = Font(name="Arial", bold=True, size=10, color="6B2D6B")
                cell.fill = PatternFill("solid", fgColor=bg)
            else:
                cell.fill = PatternFill("solid", fgColor=bg)

    # Bottom border on last data row
    last_row = 5 + max(len(docs), 1)
    for ci in range(1, len(HEADERS) + 1):
        cl   = get_column_letter(ci)
        cell = ws[f"{cl}{last_row}"]
        cell.border = Border(
            top=cell.border.top, bottom=thick,
            left=thick if ci == 1 else thin,
            right=thick if ci == len(HEADERS) else thin,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"machine_downtime_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



# ═══════════════════════════════════════════════════════════════
# MDB DAILY READINGS
# ═══════════════════════════════════════════════════════════════

@api.post("/mdb-readings")
async def create_mdb_reading(
    panel_number: str = Form(...),
    l1: float = Form(...),
    l2: float = Form(...),
    l3: float = Form(...),
    neutral: float = Form(...),
    notes: str = Form(""),
    image: Optional[UploadFile] = File(None),
    user: dict = Depends(get_current_user),
):
    image_url = None
    if image and image.filename:
        if not (image.content_type or "").startswith("image/"):
            raise HTTPException(400, "Only image files are allowed")
        contents = await image.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(400, "Image too large (max 5MB)")
        ext = Path(image.filename).suffix.lower() or ".jpg"
        fname = f"{uuid.uuid4()}{ext}"
        fpath = MDB_UPLOADS_DIR / fname
        with open(fpath, "wb") as f:
            f.write(contents)
        image_url = f"/uploads/mdb/{fname}"
    doc = {
        "id": str(uuid.uuid4()),
        "panel_number": panel_number.upper().strip(),
        "l1": l1, "l2": l2, "l3": l3, "neutral": neutral,
        "notes": notes or "",
        "image_url": image_url,
        "technician_id": user["id"],
        "technician_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.mdb_readings.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/mdb-readings/export/excel")
async def export_mdb_excel(
    request: Request,
    panel_number: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    _=Depends(require_admin),
):
    _rl_check_export(request.client.host)
    q: dict = {}
    if panel_number:
        q["panel_number"] = {"$regex": re.escape(panel_number), "$options": "i"}
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    docs = await db.mdb_readings.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)

    PURPLE       = "6B2D6B"
    PURPLE_LIGHT = "F3E8F3"
    WHITE        = "FFFFFF"
    HEADER_BG    = "4A1442"
    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color=PURPLE)

    HEADERS    = ["#", "Date", "Time", "Panel", "L1 (A)", "L2 (A)", "L3 (A)", "Neutral (A)", "Technician", "Notes"]
    COL_WIDTHS = [5,   14,     11,     14,      12,       12,       12,       14,             22,            30]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MDB Readings"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = get_column_letter(len(HEADERS))

    ws.row_dimensions[1].height = 8
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=WHITE)

    ws.row_dimensions[2].height = 80
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].fill = PatternFill("solid", fgColor=WHITE)
    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width  = sum(COL_WIDTHS) * 7
        img.height = 82
        img.anchor = "A2"
        ws.add_image(img)

    ws.row_dimensions[3].height = 38
    ws.merge_cells(f"A3:{last_col}3")
    tc = ws["A3"]
    tc.value = "MDB Daily Reading Report"
    tc.font  = Font(name="Arial", bold=True, size=16, color=PURPLE)
    tc.fill  = PatternFill("solid", fgColor=WHITE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = Border(bottom=Side(style="medium", color=PURPLE))

    ws.row_dimensions[4].height = 28
    ws.merge_cells(f"A4:{last_col}4")
    period = ""
    if date_from and date_to: period = f"   |   Period: {date_from}  to  {date_to}"
    elif date_from:            period = f"   |   From: {date_from}"
    ic = ws["A4"]
    ic.value = f"   Total Records: {len(docs)}{period}"
    ic.font  = Font(name="Arial", bold=True, size=12, color=WHITE)
    ic.fill  = PatternFill("solid", fgColor=HEADER_BG)
    ic.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[5].height = 30
    for ci, hdr in enumerate(HEADERS, 1):
        cl   = get_column_letter(ci)
        cell = ws[f"{cl}5"]
        cell.value = hdr
        cell.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=thick, bottom=thick,
            left=thick if ci == 1 else thin,
            right=thick if ci == len(HEADERS) else thin,
        )

    for ri, doc in enumerate(docs):
        row = 6 + ri
        ws.row_dimensions[row].height = 22
        bg = PURPLE_LIGHT if ri % 2 == 0 else WHITE
        dt = doc.get("created_at", "")
        values = [
            ri + 1, dt[:10], dt[11:16] if len(dt) > 10 else "",
            doc.get("panel_number", ""),
            doc.get("l1", ""), doc.get("l2", ""), doc.get("l3", ""), doc.get("neutral", ""),
            doc.get("technician_name", ""),
            doc.get("notes", "") or "",
        ]
        for ci, val in enumerate(values, 1):
            cl   = get_column_letter(ci)
            cell = ws[f"{cl}{row}"]
            cell.value = val
            cell.font  = Font(name="Arial", size=10)
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="center" if ci <= 8 else "left",
                vertical="center",
            )
            cell.border = Border(
                top=thin, bottom=thin,
                left=thick if ci == 1 else thin,
                right=thick if ci == len(HEADERS) else thin,
            )
            if ci in (5, 6, 7, 8) and isinstance(val, (int, float)):
                cell.font = Font(name="Arial", bold=True, size=10, color=PURPLE)

    last_row = 5 + max(len(docs), 1)
    for ci in range(1, len(HEADERS) + 1):
        cl   = get_column_letter(ci)
        cell = ws[f"{cl}{last_row}"]
        cell.border = Border(
            top=cell.border.top, bottom=thick,
            left=thick if ci == 1 else thin,
            right=thick if ci == len(HEADERS) else thin,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mdb_readings_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.get("/mdb-readings/export/pdf")
async def export_mdb_pdf(
    panel_number: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    _=Depends(require_admin),
):
    q: dict = {}
    if panel_number:
        q["panel_number"] = {"$regex": re.escape(panel_number), "$options": "i"}
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    docs = await db.mdb_readings.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)

    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    pdfdoc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    PURPLE_C = rl_colors.Color(0x6B/255, 0x2D/255, 0x6B/255)
    HEADER_C = rl_colors.Color(0x4A/255, 0x14/255, 0x42/255)
    LIGHT_C  = rl_colors.Color(0xF3/255, 0xE8/255, 0xF3/255)

    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("t", parent=styles["Heading1"],
                               textColor=PURPLE_C, fontSize=15, spaceAfter=4)
    sub_st   = ParagraphStyle("s", parent=styles["Normal"],
                               textColor=rl_colors.grey, fontSize=9, spaceAfter=10)
    elements = []

    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")
    if os.path.exists(LOGO_PATH):
        elements.append(RLImage(LOGO_PATH, width=6*cm, height=2*cm))
        elements.append(Spacer(1, 0.3*cm))

    elements.append(Paragraph("MDB Daily Reading Report", title_st))
    period_str = ""
    if date_from and date_to: period_str = f" | Period: {date_from} to {date_to}"
    elif date_from:            period_str = f" | From: {date_from}"
    elements.append(Paragraph(f"Total Records: {len(docs)}{period_str}", sub_st))

    hdrs = ["#", "Date", "Time", "Panel", "L1 (A)", "L2 (A)", "L3 (A)", "N (A)", "Technician", "Notes"]
    table_data = [hdrs]
    for i, d in enumerate(docs, 1):
        dt = d.get("created_at", "")
        table_data.append([
            str(i), dt[:10], dt[11:16] if len(dt) > 10 else "",
            d.get("panel_number", ""),
            str(d.get("l1", "")), str(d.get("l2", "")),
            str(d.get("l3", "")), str(d.get("neutral", "")),
            d.get("technician_name", ""),
            d.get("notes", "") or "",
        ])

    col_w = [1*cm, 2.5*cm, 2*cm, 2.5*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 4*cm, 5*cm]
    tbl = Table(table_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  HEADER_C),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  rl_colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  9),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",         (9, 1), (9, -1),  "LEFT"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [rl_colors.white, LIGHT_C]),
        ("GRID",          (0, 0), (-1, -1), 0.5, rl_colors.Color(0xCC/255, 0xCC/255, 0xCC/255)),
        ("BOX",           (0, 0), (-1, -1), 1.5, PURPLE_C),
        ("FONTNAME",      (4, 1), (7, -1),  "Helvetica-Bold"),
        ("TEXTCOLOR",     (4, 1), (7, -1),  PURPLE_C),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)
    pdfdoc.build(elements)

    buf.seek(0)
    fname = f"mdb_readings_{date_from or 'all'}_{date_to or 'all'}.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.get("/mdb-readings")
async def list_mdb_readings(
    panel_number: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    q: dict = {}
    if panel_number:
        q["panel_number"] = {"$regex": re.escape(panel_number), "$options": "i"}
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    docs = await db.mdb_readings.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs


@api.delete("/mdb-readings/{rid}")
async def delete_mdb_reading(rid: str, _=Depends(require_admin)):
    doc = await db.mdb_readings.find_one({"id": rid})
    if not doc:
        raise HTTPException(404, "Not found")
    if doc.get("image_url"):
        try:
            fpath = ROOT_DIR / doc["image_url"].lstrip("/")
            if fpath.exists():
                fpath.unlink()
        except Exception:
            pass
    await db.mdb_readings.delete_one({"id": rid})
    return {"ok": True}



# ═══════════════════════════════════════════════════════════════
# MAINTENANCE ANALYTICS
# ═══════════════════════════════════════════════════════════════

def _calc_duration_minutes(start_str: str, end_str: str) -> int:
    """Return duration in minutes between two 12-hour time strings, or 0."""
    import re as _re
    def _p(s):
        if not s: return None
        m = _re.match(r'^(\d{1,2}):(\d{2})\s*(AM|PM)$', s.strip(), _re.IGNORECASE)
        if not m: return None
        h, mn, ap = int(m.group(1)), int(m.group(2)), m.group(3).upper()
        if ap == 'PM' and h != 12: h += 12
        if ap == 'AM' and h == 12: h = 0
        return h * 60 + mn
    s, e = _p(start_str), _p(end_str)
    if s is None or e is None: return 0
    diff = e - s
    if diff < 0: diff += 1440
    return diff


@api.get("/analytics/maintenance")
async def get_maintenance_analytics(
    year: Optional[int] = None,
    month: Optional[int] = None,
    machine_id: Optional[str] = None,
    specialty: Optional[str] = None,
    _=Depends(require_admin),
):
    now = datetime.now(timezone.utc)
    eff_year = year or now.year

    # Build query for the filtered period
    q: dict = {}
    if year and month:
        nx_y, nx_m = (eff_year + 1, 1) if month == 12 else (eff_year, month + 1)
        q["created_at"] = {"$gte": f"{eff_year:04d}-{month:02d}-01", "$lt": f"{nx_y:04d}-{nx_m:02d}-01"}
    elif year:
        q["created_at"] = {"$gte": f"{eff_year:04d}-01-01", "$lt": f"{eff_year+1:04d}-01-01"}
    if machine_id:
        q["machine_id"] = machine_id

    all_docs = await db.breakdowns.find(q, {"_id": 0}).sort("created_at", 1).to_list(5000)

    # Resolve technician specialties
    tech_ids = list({d["technician_id"] for d in all_docs if d.get("technician_id")})
    users_cur = await db.users.find({"id": {"$in": tech_ids}}, {"_id": 0, "id": 1, "specialty": 1}).to_list(500)
    tech_specialty = {u["id"]: u.get("specialty") for u in users_cur}

    if specialty:
        all_docs = [d for d in all_docs if tech_specialty.get(d.get("technician_id", "")) == specialty]

    for d in all_docs:
        d["_minutes"] = _calc_duration_minutes(d.get("start_time", ""), d.get("end_time", ""))

    total_breakdowns = len(all_docs)
    total_minutes = sum(d["_minutes"] for d in all_docs)
    has_dur = [d for d in all_docs if d["_minutes"] > 0]
    mttr_min = (sum(d["_minutes"] for d in has_dur) / len(has_dur)) if has_dur else 0

    # MTBF estimate
    if total_breakdowns > 1:
        try:
            dt1 = datetime.fromisoformat(all_docs[0]["created_at"].replace("Z", "+00:00"))
            dt2 = datetime.fromisoformat(all_docs[-1]["created_at"].replace("Z", "+00:00"))
            period_h = max((dt2 - dt1).total_seconds() / 3600, 1.0)
        except Exception:
            period_h = 720.0
        mtbf_h = (period_h - total_minutes / 60) / total_breakdowns
    else:
        mtbf_h = 0.0

    elec_docs = [d for d in all_docs if tech_specialty.get(d.get("technician_id", "")) == "electrical"]
    mech_docs = [d for d in all_docs if tech_specialty.get(d.get("technician_id", "")) == "mechanical"]

    machine_stats: dict = defaultdict(lambda: {"count": 0, "minutes": 0, "name": ""})
    reason_counts: dict = defaultdict(int)
    for d in all_docs:
        key = d.get("machine_number", "Unknown")
        machine_stats[key]["count"] += 1
        machine_stats[key]["minutes"] += d["_minutes"]
        machine_stats[key]["name"] = d.get("machine_name", "")
        reason_counts[d.get("brief_description", "Other")] += 1

    most_failed = max(machine_stats, key=lambda k: machine_stats[k]["count"]) if machine_stats else "—"

    by_machine = sorted(
        [{"machine": k, "name": v["name"], "count": v["count"],
          "downtime_hours": round(v["minutes"] / 60, 1)}
         for k, v in machine_stats.items()],
        key=lambda x: x["count"], reverse=True,
    )[:10]

    top_reasons = sorted(
        [{"reason": k, "count": v} for k, v in reason_counts.items()],
        key=lambda x: x["count"], reverse=True,
    )[:8]

    # Preventive maintenance count for the filtered period
    pm_q: dict = {"category": "preventive"}
    if q.get("created_at"):
        pm_q["created_at"] = q["created_at"]
    pm_total = await db.inspections.count_documents(pm_q)

    # Monthly trend — last 12 months (2 DB calls instead of 24)
    MNAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    trend_start_m = now.month - 11
    trend_start_y = now.year
    while trend_start_m <= 0:
        trend_start_m += 12
        trend_start_y -= 1
    trend_start_str = f"{trend_start_y:04d}-{trend_start_m:02d}-01"

    trend_bdocs = await db.breakdowns.find(
        {"created_at": {"$gte": trend_start_str}},
        {"_id": 0, "created_at": 1, "start_time": 1, "end_time": 1},
    ).to_list(10000)
    bd_by_ym: dict = defaultdict(list)
    for d in trend_bdocs:
        bd_by_ym[d["created_at"][:7]].append(d)

    pm_agg = await db.inspections.aggregate([
        {"$match": {"category": "preventive", "created_at": {"$gte": trend_start_str}}},
        {"$group": {"_id": {"$substr": ["$created_at", 0, 7]}, "count": {"$sum": 1}}},
    ]).to_list(50)
    pm_by_ym = {r["_id"]: r["count"] for r in pm_agg}

    trend = []
    for i in range(11, -1, -1):
        mo = now.month - i
        yr = now.year
        while mo <= 0:
            mo += 12
            yr -= 1
        ym_key = f"{yr:04d}-{mo:02d}"
        month_docs = bd_by_ym.get(ym_key, [])
        trend.append({
            "month": f"{MNAMES[mo-1]} {yr}",
            "count": len(month_docs),
            "downtime_hours": round(sum(_calc_duration_minutes(d.get("start_time",""), d.get("end_time","")) for d in month_docs) / 60, 1),
            "pm_count": pm_by_ym.get(ym_key, 0),
        })

    mttr_h = mttr_min / 60
    mtbf_h_safe = max(mtbf_h, 0)
    availability_pct = round((mtbf_h_safe / (mtbf_h_safe + mttr_h)) * 100, 1) if (mtbf_h_safe + mttr_h) > 0 else 0.0

    return {
        "overview": {
            "total_breakdowns": total_breakdowns,
            "total_downtime_hours": round(total_minutes / 60, 1),
            "mttr_minutes": round(mttr_min, 1),
            "mtbf_hours": round(mtbf_h_safe, 1),
            "availability_pct": availability_pct,
            "electrical_count": len(elec_docs),
            "mechanical_count": len(mech_docs),
            "most_failed_machine": most_failed,
            "pm_count": pm_total,
        },
        "monthly_trend": trend,
        "by_machine": by_machine,
        "top_reasons": top_reasons,
        "by_specialty": {
            "electrical": {"count": len(elec_docs), "downtime_hours": round(sum(d["_minutes"] for d in elec_docs)/60,1)},
            "mechanical": {"count": len(mech_docs), "downtime_hours": round(sum(d["_minutes"] for d in mech_docs)/60,1)},
        },
    }


@api.get("/analytics/maintenance/export/excel")
async def export_analytics_excel(
    request: Request,
    year: Optional[int] = None,
    month: Optional[int] = None,
    machine_id: Optional[str] = None,
    specialty: Optional[str] = None,
    _=Depends(require_admin),
):
    _rl_check_export(request.client.host)
    now = datetime.now(timezone.utc)
    eff_year = year or now.year
    q: dict = {}
    if year and month:
        nx_y, nx_m = (eff_year + 1, 1) if month == 12 else (eff_year, month + 1)
        q["created_at"] = {"$gte": f"{eff_year:04d}-{month:02d}-01", "$lt": f"{nx_y:04d}-{nx_m:02d}-01"}
    elif year:
        q["created_at"] = {"$gte": f"{eff_year:04d}-01-01", "$lt": f"{eff_year+1:04d}-01-01"}
    if machine_id:
        q["machine_id"] = machine_id

    docs = await db.breakdowns.find(q, {"_id": 0}).sort("created_at", 1).to_list(5000)
    tech_ids = list({d["technician_id"] for d in docs if d.get("technician_id")})
    users_cur = await db.users.find({"id": {"$in": tech_ids}}, {"_id": 0, "id": 1, "specialty": 1}).to_list(500)
    tech_specialty = {u["id"]: u.get("specialty") for u in users_cur}
    if specialty:
        docs = [d for d in docs if tech_specialty.get(d.get("technician_id","")) == specialty]
    for d in docs:
        d["_specialty"] = tech_specialty.get(d.get("technician_id",""), "")
        d["_minutes"]   = _calc_duration_minutes(d.get("start_time",""), d.get("end_time",""))

    PURPLE       = "6B2D6B"
    PURPLE_LIGHT = "F3E8F3"
    WHITE        = "FFFFFF"
    HEADER_BG    = "4A1442"
    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color=PURPLE)

    HEADERS    = ["#", "Date", "Machine", "Specialty", "Reason", "Technician", "Start", "End", "Duration (min)", "Status"]
    COL_WIDTHS = [5,   16,     18,        14,           32,       20,           12,      12,    16,               12]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Analytics"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = get_column_letter(len(HEADERS))

    ws.row_dimensions[1].height = 8
    for c in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=WHITE)

    ws.row_dimensions[2].height = 80
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].fill = PatternFill("solid", fgColor=WHITE)
    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = sum(COL_WIDTHS) * 7
        img.height = 82
        img.anchor = "A2"
        ws.add_image(img)

    ws.row_dimensions[3].height = 38
    ws.merge_cells(f"A3:{last_col}3")
    tc = ws["A3"]
    period_lbl = f" — {eff_year}" + (f"/{month:02d}" if month else "")
    tc.value = f"Maintenance Analytics Report{period_lbl}"
    tc.font  = Font(name="Arial", bold=True, size=16, color=PURPLE)
    tc.fill  = PatternFill("solid", fgColor=WHITE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = Border(bottom=Side(style="medium", color=PURPLE))

    ws.row_dimensions[4].height = 28
    ws.merge_cells(f"A4:{last_col}4")
    ic = ws["A4"]
    ic.value = f"   Total Records: {len(docs)}   |   Specialty: {specialty or 'All'}"
    ic.font  = Font(name="Arial", bold=True, size=12, color=WHITE)
    ic.fill  = PatternFill("solid", fgColor=HEADER_BG)
    ic.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[5].height = 30
    for ci, hdr in enumerate(HEADERS, 1):
        cell = ws[f"{get_column_letter(ci)}5"]
        cell.value = hdr
        cell.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thick, bottom=thick,
                             left=thick if ci==1 else thin,
                             right=thick if ci==len(HEADERS) else thin)

    for ri, doc in enumerate(docs):
        row = 6 + ri
        ws.row_dimensions[row].height = 22
        bg   = PURPLE_LIGHT if ri % 2 == 0 else WHITE
        vals = [
            ri + 1,
            doc.get("created_at","")[:10],
            doc.get("machine_number",""),
            doc.get("_specialty","").capitalize() or "—",
            doc.get("brief_description",""),
            doc.get("technician_name",""),
            doc.get("start_time","") or "—",
            doc.get("end_time","")   or "—",
            doc["_minutes"] if doc["_minutes"] > 0 else "—",
            "Resolved" if doc.get("status") == "resolved" else "Open",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws[f"{get_column_letter(ci)}{row}"]
            cell.value = val
            cell.font  = Font(name="Arial", size=10)
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if ci in (1,9,10) else "left", vertical="center")
            cell.border = Border(top=thin, bottom=thin,
                                 left=thick if ci==1 else thin,
                                 right=thick if ci==len(HEADERS) else thin)

    last_row = 5 + max(len(docs), 1)
    for ci in range(1, len(HEADERS)+1):
        cell = ws[f"{get_column_letter(ci)}{last_row}"]
        cell.border = Border(top=cell.border.top, bottom=thick,
                             left=thick if ci==1 else thin,
                             right=thick if ci==len(HEADERS) else thin)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"analytics_{eff_year}_{month or 'all'}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.get("/analytics/maintenance/export/pdf")
async def export_analytics_pdf(
    year: Optional[int] = None,
    month: Optional[int] = None,
    machine_id: Optional[str] = None,
    specialty: Optional[str] = None,
    _=Depends(require_admin),
):
    now = datetime.now(timezone.utc)
    eff_year = year or now.year
    q: dict = {}
    if year and month:
        nx_y, nx_m = (eff_year + 1, 1) if month == 12 else (eff_year, month + 1)
        q["created_at"] = {"$gte": f"{eff_year:04d}-{month:02d}-01", "$lt": f"{nx_y:04d}-{nx_m:02d}-01"}
    elif year:
        q["created_at"] = {"$gte": f"{eff_year:04d}-01-01", "$lt": f"{eff_year+1:04d}-01-01"}
    if machine_id:
        q["machine_id"] = machine_id

    docs = await db.breakdowns.find(q, {"_id": 0}).sort("created_at", 1).to_list(5000)
    tech_ids = list({d["technician_id"] for d in docs if d.get("technician_id")})
    users_cur = await db.users.find({"id": {"$in": tech_ids}}, {"_id": 0, "id": 1, "specialty": 1}).to_list(500)
    tech_specialty = {u["id"]: u.get("specialty") for u in users_cur}
    if specialty:
        docs = [d for d in docs if tech_specialty.get(d.get("technician_id","")) == specialty]
    for d in docs:
        d["_specialty"] = tech_specialty.get(d.get("technician_id",""), "")
        d["_minutes"]   = _calc_duration_minutes(d.get("start_time",""), d.get("end_time",""))

    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    pdfdoc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    PURPLE_C = rl_colors.Color(0x6B/255, 0x2D/255, 0x6B/255)
    HEADER_C = rl_colors.Color(0x4A/255, 0x14/255, 0x42/255)
    LIGHT_C  = rl_colors.Color(0xF3/255, 0xE8/255, 0xF3/255)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("t", parent=styles["Heading1"], textColor=PURPLE_C, fontSize=15, spaceAfter=4)
    sub_st   = ParagraphStyle("s", parent=styles["Normal"], textColor=rl_colors.grey, fontSize=9, spaceAfter=10)
    elements = []
    LOGO_PATH = os.path.join(ROOT_DIR, "company_logo.jpeg")
    if os.path.exists(LOGO_PATH):
        elements.append(RLImage(LOGO_PATH, width=6*cm, height=2*cm))
        elements.append(Spacer(1, 0.3*cm))
    period_lbl = f"{eff_year}" + (f"/{month:02d}" if month else "")
    elements.append(Paragraph(f"Maintenance Analytics Report — {period_lbl}", title_st))
    elements.append(Paragraph(f"Total Records: {len(docs)}   |   Specialty: {specialty or 'All'}", sub_st))

    hdrs = ["#", "Date", "Machine", "Specialty", "Reason", "Technician", "Duration (min)", "Status"]
    table_data = [hdrs]
    for i, d in enumerate(docs, 1):
        table_data.append([
            str(i),
            d.get("created_at","")[:10],
            d.get("machine_number",""),
            (d.get("_specialty","") or "—").capitalize(),
            d.get("brief_description",""),
            d.get("technician_name",""),
            str(d["_minutes"]) if d["_minutes"] > 0 else "—",
            "Resolved" if d.get("status") == "resolved" else "Open",
        ])
    col_w = [1*cm, 2.5*cm, 3*cm, 2.5*cm, 6*cm, 4*cm, 3*cm, 2.5*cm]
    tbl = Table(table_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  HEADER_C),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  rl_colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  9),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",         (4, 1), (5, -1),  "LEFT"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [rl_colors.white, LIGHT_C]),
        ("GRID",          (0, 0), (-1, -1), 0.5, rl_colors.Color(0xCC/255,0xCC/255,0xCC/255)),
        ("BOX",           (0, 0), (-1, -1), 1.5, PURPLE_C),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)
    pdfdoc.build(elements)
    buf.seek(0)
    fname = f"analytics_{eff_year}_{month or 'all'}.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ═══════════════════════════════════════════════════════════════
# FORM PERMISSIONS
# ═══════════════════════════════════════════════════════════════

FORM_IDS = {
    "inspection_electrical", "inspection_panels",
    "inspection_mechanical", "inspection_chiller",
    "inspection_cooling_tower", "inspection_preventive",
    "breakdown", "mdb_reading",
}

SPECIALTY_DEFAULTS = {
    "electrical": {
        "inspection_electrical": True,  "inspection_panels": True,
        "inspection_mechanical": False, "inspection_chiller": False,
        "inspection_cooling_tower": False, "inspection_preventive": False,
        "breakdown": True, "mdb_reading": True,
    },
    "mechanical": {
        "inspection_electrical": False, "inspection_panels": False,
        "inspection_mechanical": True,  "inspection_chiller": True,
        "inspection_cooling_tower": True, "inspection_preventive": False,
        "breakdown": True, "mdb_reading": True,
    },
}

class FormPermissionBody(BaseModel):
    forms: dict  # {form_id: bool}

@api.get("/form-permissions")
async def list_form_permissions(_=Depends(require_admin)):
    docs = await db.form_permissions.find({}, {"_id": 0}).to_list(500)
    return docs

@api.put("/form-permissions/specialty/{specialty}")
async def set_specialty_permissions(specialty: str, body: FormPermissionBody, _=Depends(require_admin)):
    if specialty not in ("electrical", "mechanical"):
        raise HTTPException(400, "Invalid specialty")
    forms = {k: bool(v) for k, v in body.forms.items() if k in FORM_IDS}
    await db.form_permissions.update_one(
        {"scope": "specialty", "target": specialty},
        {"$set": {"scope": "specialty", "target": specialty, "forms": forms,
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return {"ok": True}

@api.put("/form-permissions/user/{user_id}")
async def set_user_permissions(user_id: str, body: FormPermissionBody, _=Depends(require_admin)):
    user_doc = await db.users.find_one({"id": user_id})
    if not user_doc:
        raise HTTPException(404, "User not found")
    forms = {k: bool(v) for k, v in body.forms.items() if k in FORM_IDS}
    await db.form_permissions.update_one(
        {"scope": "user", "target": user_id},
        {"$set": {"scope": "user", "target": user_id, "forms": forms,
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return {"ok": True}

@api.delete("/form-permissions/user/{user_id}")
async def delete_user_permissions(user_id: str, _=Depends(require_admin)):
    await db.form_permissions.delete_one({"scope": "user", "target": user_id})
    return {"ok": True}

@api.get("/form-permissions/me")
async def get_my_permissions(user=Depends(get_current_user)):
    if user["role"] == "admin":
        return {f: True for f in FORM_IDS}
    spec = user.get("specialty", "")
    # User-specific override first
    doc = await db.form_permissions.find_one({"scope": "user", "target": user["id"]})
    if doc:
        forms = doc.get("forms", {})
        base = SPECIALTY_DEFAULTS.get(spec, {f: True for f in FORM_IDS})
        return {f: forms.get(f, base.get(f, True)) for f in FORM_IDS}
    # Specialty DB override
    doc = await db.form_permissions.find_one({"scope": "specialty", "target": spec})
    if doc:
        forms = doc.get("forms", {})
        base = SPECIALTY_DEFAULTS.get(spec, {f: True for f in FORM_IDS})
        return {f: forms.get(f, base.get(f, True)) for f in FORM_IDS}
    # Specialty hardcoded defaults
    return SPECIALTY_DEFAULTS.get(spec, {f: True for f in FORM_IDS})


@api.get("/stats/monthly-breakdown")
async def stats_monthly_breakdown(_=Depends(require_admin)):
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    docs = await db.breakdowns.find({"created_at": {"$gte": month_start}}, {"_id": 0}).to_list(1000)
    total = len(docs)
    resolved = sum(1 for d in docs if d.get("status") == "resolved")
    cause_counts: dict = defaultdict(int)
    total_minutes = 0
    for d in docs:
        cause_counts[d.get("brief_description", "—")[:60]] += 1
        try:
            if d.get("start_time") and d.get("end_time"):
                s = datetime.fromisoformat(d["start_time"].replace("Z", "+00:00"))
                e = datetime.fromisoformat(d["end_time"].replace("Z", "+00:00"))
                diff = (e - s).total_seconds() / 60
                if 0 < diff < 10000:
                    total_minutes += diff
        except Exception:
            pass
    top_causes = sorted(
        [{"cause": k, "count": v} for k, v in cause_counts.items()],
        key=lambda x: -x["count"],
    )[:5]
    return {
        "total": total,
        "resolved": resolved,
        "open": total - resolved,
        "downtime_hours": round(total_minutes / 60, 1),
        "top_causes": top_causes,
    }

@api.get("/stats/overview")
async def stats_overview(_=Depends(require_admin)):
    total_inspections = await db.inspections.count_documents({})
    total_machines = await db.machines.count_documents({})
    total_chillers = await db.chillers.count_documents({})
    total_panels = await db.panels.count_documents({})
    total_techs = await db.users.count_documents({"role": "technician"})
    total_questions = await db.questions.count_documents({})
    today = datetime.now(timezone.utc).date().isoformat()
    today_count = await db.inspections.count_documents({"created_at": {"$gte": today}})
    pipeline = [{"$unwind": "$answers"}, {"$match": {"answers.a": False}}, {"$count": "fails"}]
    fails = 0
    async for d in db.inspections.aggregate(pipeline):
        fails = d["fails"]
    open_fails = await _count_open_failures()
    return {"total_inspections": total_inspections, "total_machines": total_machines,
            "total_chillers": total_chillers, "total_panels": total_panels,
            "total_technicians": total_techs, "total_questions": total_questions,
            "today_inspections": today_count, "total_fails": fails,
            "open_fails": open_fails}


# ---------- Notifications ----------
@api.get("/notifications")
async def list_notifications(_=Depends(require_admin)):
    items = await db.notifications.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    unread = sum(1 for n in items if not n.get("is_read"))
    return {"items": items, "unread": unread}


@api.patch("/notifications/{nid}/read")
async def mark_notification_read(nid: str, _=Depends(require_admin)):
    await db.notifications.update_one({"id": nid}, {"$set": {"is_read": True}})
    return {"ok": True}


@api.patch("/notifications/read-all")
async def mark_all_read(_=Depends(require_admin)):
    await db.notifications.update_many({"is_read": False}, {"$set": {"is_read": True}})
    return {"ok": True}


@api.delete("/notifications/{nid}")
async def delete_notification(nid: str, _=Depends(require_admin)):
    await db.notifications.delete_one({"id": nid})
    return {"ok": True}


# ---------- Audit Log ----------
@api.get("/audit-logs")
async def list_audit_logs(
    page: int = 1,
    limit: int = 50,
    resource_type: Optional[str] = None,
    actor_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    _=Depends(require_admin),
):
    q: dict = {}
    if resource_type: q["resource_type"] = resource_type
    if actor_id: q["actor_id"] = actor_id
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    skip = (page - 1) * limit
    total = await db.audit_logs.count_documents(q)
    items = await db.audit_logs.find(q, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"total": total, "page": page, "items": items}


# ---------- Seed ----------
DEFAULT_QUESTIONS = {
    "electrical": ["فحص لوحة الكهرباء الرئيسية", "التحقق من الكابلات والتوصيلات",
                   "قياس الجهد الكهربائي", "فحص الفيوزات والقواطع", "التأكد من التأريض"],
    "mechanical": ["فحص المحاور والبيرنقات", "التحقق من مستوى الزيت",
                   "فحص الأحزمة والسيور", "فحص البراغي والمسامير", "التأكد من عمل الصمامات"],
    "chiller": ["فحص درجة حرارة المبرد", "التحقق من ضغط الفريون",
                "فحص مروحة المكثف", "تنظيف فلاتر الشيلر", "فحص تسرب سوائل التبريد"],
    "panels": ["فحص مفاتيح اللوحة", "فحص المؤشرات والإشارات", "قياس درجة حرارة اللوحة",
               "فحص توصيلات الـ PLC", "التأكد من إغلاق اللوحة بإحكام"],
}

DEFAULT_TECHS = [
    ("EMP-001", "محمد", "tech123", "electrical"),
    ("EMP-002", "أحمد", "tech123", "electrical"),
    ("EMP-003", "خالد", "tech123", "mechanical"),
    ("EMP-004", "عبدالله", "tech123", "mechanical"),
]
ADMIN_EMP_NUMBER = "ADMIN-001"


@app.on_event("startup")
async def startup():
    # Migration: drop legacy email unique index if present
    try:
        existing_indexes = await db.users.index_information()
        if "email_1" in existing_indexes:
            await db.users.drop_index("email_1")
            logger.info("Dropped legacy users.email_1 index")
    except Exception as e:
        logger.warning(f"Index check failed (non-fatal): {e}")

    # Migration: backfill employee_number for any legacy user
    legacy = await db.users.find({"employee_number": {"$exists": False}}, {"_id": 0}).to_list(500)
    for idx, u in enumerate(legacy):
        if u.get("role") == "admin":
            emp = ADMIN_EMP_NUMBER
        else:
            # reuse email prefix if present, else sequential
            email = u.get("email") or ""
            prefix = email.split("@")[0].upper() if "@" in email else f"EMP-{idx+1:03d}"
            emp = prefix if prefix.startswith("EMP") or prefix.startswith("TECH") else f"EMP-{idx+1:03d}"
        # ensure unique
        base = emp
        suffix = 1
        while await db.users.find_one({"employee_number": emp, "id": {"$ne": u["id"]}}):
            suffix += 1
            emp = f"{base}-{suffix}"
        await db.users.update_one(
            {"id": u["id"]},
            {"$set": {"employee_number": emp}, "$unset": {"email": ""}},
        )

    await db.users.create_index("employee_number", unique=True)
    await db.users.create_index("id", unique=True)
    for c in ("machines", "chillers", "panels", "cooling_towers"):
        await db[c].create_index("number", unique=True)
        await db[c].create_index("id", unique=True)
    await db.questions.create_index("id", unique=True)
    await db.questions.create_index("category")
    await db.inspections.create_index("id", unique=True)
    await db.inspections.create_index("created_at")
    await db.inspections.create_index("target_id")
    await db.inspections.create_index("category")
    await db.inspections.create_index([("target_id", 1), ("category", 1)])
    await db.inspections.create_index([("category", 1), ("created_at", -1)])
    await db.inspections.create_index([("target_number", 1), ("created_at", -1)])
    await db.inspections.create_index("target_number")
    await db.breakdowns.create_index("id", unique=True)
    await db.breakdowns.create_index("created_at")
    await db.breakdowns.create_index("machine_id")
    await db.breakdowns.create_index([("created_at", 1), ("machine_id", 1)])
    await db.mdb_readings.create_index("created_at")
    for c in ("machines", "chillers", "panels", "cooling_towers"):
        await db[c].create_index("sort_order")
    await db.registration_requests.create_index("status")
    await db.registration_requests.create_index("expires_at", expireAfterSeconds=0, sparse=True)
    await db.audit_logs.create_index("created_at", expireAfterSeconds=90 * 24 * 3600)
    await db.audit_logs.create_index("actor_id")
    await db.audit_logs.create_index([("resource_type", 1), ("created_at", -1)])
    await db.notifications.create_index("created_at", expireAfterSeconds=30 * 24 * 3600)
    await db.notifications.create_index("is_read")

    # Migration: backfill sort_order for chillers, panels, cooling_towers missing it
    for coll_name in ("chillers", "panels", "cooling_towers"):
        items = await db[coll_name].find(
            {"sort_order": {"$exists": False}}, {"_id": 0, "id": 1}
        ).sort("number", 1).to_list(2000)
        for i, item in enumerate(items):
            await db[coll_name].update_one({"id": item["id"]}, {"$set": {"sort_order": i}})
        if items:
            logger.info(f"Backfilled sort_order for {len(items)} {coll_name}")

    # Migration: convert inspection answers from long to short field names
    migrated = 0
    async for doc in db.inspections.find(
        {"answers.0.question_id": {"$exists": True}}, {"_id": 1, "answers": 1}
    ):
        new_ans = []
        for a in doc.get("answers", []):
            na: dict = {"qid": a["question_id"], "a": a.get("answer")}
            if a.get("numeric_value") is not None:
                na["nv"] = a["numeric_value"]
            if a.get("note"):
                na["n"] = a["note"]
            new_ans.append(na)
        await db.inspections.update_one({"_id": doc["_id"]}, {"$set": {"answers": new_ans}})
        migrated += 1
    if migrated:
        logger.info(f"Migrated {migrated} inspections to short answer field names")

    # Seed chiller questions if none exist
    if await db.questions.count_documents({"category": "chiller"}) == 0:
        now = datetime.now(timezone.utc).isoformat()
        chiller_qs = [
            {"text": "Check temperature setting (must be 8–25°C)",          "answer_type": "numeric", "unit": "°C"},
            {"text": "Check actual temperature – High",                      "answer_type": "numeric", "unit": "°C"},
            {"text": "Check actual temperature – Low",                       "answer_type": "numeric", "unit": "°C"},
            {"text": "Check water pump pressure – High",                     "answer_type": "numeric", "unit": "kg"},
            {"text": "Check water pump pressure – Low",                      "answer_type": "numeric", "unit": "kg"},
            {"text": "Check any abnormal sound in the unit",                 "answer_type": "yes_no",  "unit": None},
            {"text": "Clean air filter if necessary",                        "answer_type": "yes_no",  "unit": None},
            {"text": "Check cooling fan if working (air cooled chiller)",    "answer_type": "yes_no",  "unit": None},
            {"text": "Check any water leakage",                              "answer_type": "yes_no",  "unit": None},
            {"text": "Check all water valves",                               "answer_type": "yes_no",  "unit": None},
            {"text": "Clean water quality in the tank, change if necessary", "answer_type": "yes_no",  "unit": None},
            {"text": "Fill up with water + Anti-rust Chemical",              "answer_type": "yes_no",  "unit": None},
        ]
        await db.questions.insert_many([
            {"id": str(uuid.uuid4()), "category": "chiller", "text": q["text"],
             "order": i, "answer_type": q["answer_type"], "unit": q["unit"], "created_at": now}
            for i, q in enumerate(chiller_qs)
        ])
        logger.info("Seeded 12 chiller questions")

    # Seed mechanical questions if none exist
    if await db.questions.count_documents({"category": "mechanical"}) == 0:
        now = datetime.now(timezone.utc).isoformat()
        mech_qs = [
            "Toggle bolts", "Oil leakage", "M/C saddle", "Oil temperature",
            "Tie bars", "Hydraulic oil level", "Ejector butterfly",
            "Abnormal sound or movement", "Hopper water",
        ]
        await db.questions.insert_many([
            {"id": str(uuid.uuid4()), "category": "mechanical", "text": q,
             "order": i, "answer_type": "yes_no", "unit": None, "created_at": now}
            for i, q in enumerate(mech_qs)
        ])
        logger.info("Seeded 9 mechanical questions")

    # Seed cooling tower questions if none exist
    if await db.questions.count_documents({"category": "cooling_tower"}) == 0:
        now = datetime.now(timezone.utc).isoformat()
        ct_qs = [
            "Check main tank water level",
            "Check water distribution over cooling pad",
            "Check exhaust fan operation",
            "Check main tank water supply level",
        ]
        await db.questions.insert_many([
            {"id": str(uuid.uuid4()), "category": "cooling_tower", "text": q,
             "order": i, "answer_type": "yes_no", "unit": None, "created_at": now}
            for i, q in enumerate(ct_qs)
        ])
        logger.info("Seeded 4 cooling tower questions")

    # Seed preventive maintenance questions if none exist
    if await db.questions.count_documents({"category": "preventive"}) == 0:
        now = datetime.now(timezone.utc).isoformat()
        prev_qs = [
            ("Clamp Unit", "Check hydraulic oil level"),
            ("Clamp Unit", "Check hydraulic valves condition"),
            ("Clamp Unit", "Check pipes and hoses for leakage"),
            ("Clamp Unit", "Check hydraulic oil temperature"),
            ("Clamp Unit", "Check front/rear safety doors"),
            ("Clamp Unit", "Remove old grease from guide shoes"),
            ("Clamp Unit", "Grease nipples and rods of guide shoes"),
            ("Clamp Unit", "Check moving platen guide shoes"),
            ("Clamp Unit", "Check for water/oil leakage"),
            ("Clamp Unit", "Check tie bars and chains condition"),
            ("Clamp Unit", "Check heat exchanger"),
            ("Clamp Unit", "Check ejector piston and butterfly"),
            ("Clamp Unit", "Check door safety switches"),
            ("Clamp Unit", "Check for loose/broken bolts"),
            ("Clamp Unit", "Check wiring and cables"),
            ("Clamp Unit", "Check clamp/ejector scale"),
            ("Clamp Unit", "Lubricate all moving parts"),
            ("Clamp Unit", "Check for abnormal noise"),
            ("Injection Unit", "Check injection piston cylinders"),
            ("Injection Unit", "Check oil leakage in pipes and hoses"),
            ("Injection Unit", "Check manifold valves"),
            ("Injection Unit", "Check motor and pump condition"),
            ("Injection Unit", "Check carriage bolts and nuts"),
            ("Injection Unit", "Check barrel water circulation"),
            ("Injection Unit", "Check hydraulic motor"),
            ("Injection Unit", "Check nozzle heater"),
            ("Injection Unit", "Check barrel heaters and thermocouples"),
            ("Injection Unit", "Check lubrication nipples"),
            ("Injection Unit", "Lubricate motor and moving parts"),
            ("Injection Unit", "Check purge shield"),
            ("Injection Unit", "Check injection unit cables and wiring"),
            ("Injection Unit", "Check injection level and nozzle centralization"),
            ("Electrical Panel", "Clean motor and wiring"),
            ("Electrical Panel", "Check emergency stop buttons"),
            ("Electrical Panel", "Check display and control wiring"),
            ("Electrical Panel", "Check cooling fans operation"),
            ("Electrical Panel", "Check voltage measurement"),
            ("Electrical Panel", "Check circuit breakers"),
            ("Electrical Panel", "Check contactors and relays"),
            ("Electrical Panel", "Check cables and wiring connections"),
            ("Electrical Panel", "Clean dust from electrical cabinet"),
            ("Electrical Panel", "Check battery and power supply"),
            ("Electrical Panel", "Check servo motor driver"),
            ("General", "Check machine level"),
            ("General", "Clean machine exterior"),
        ]
        await db.questions.insert_many([
            {"id": str(uuid.uuid4()), "category": "preventive", "section": sec,
             "text": txt, "order": i, "answer_type": "yes_no", "unit": None, "created_at": now}
            for i, (sec, txt) in enumerate(prev_qs)
        ])
        logger.info("Seeded 45 preventive maintenance questions")

    admin_pw = os.environ.get("ADMIN_PASSWORD", "admin123")
    admin_emp = os.environ.get("ADMIN_EMPLOYEE_NUMBER", ADMIN_EMP_NUMBER).upper()
    existing = await db.users.find_one({"employee_number": admin_emp})
    if not existing:
        await db.users.insert_one({"id": str(uuid.uuid4()), "employee_number": admin_emp,
                                   "name": "ادمن", "role": "admin", "specialty": None,
                                   "password_hash": hash_password(admin_pw),
                                   "created_at": datetime.now(timezone.utc).isoformat()})
    else:
        updates = {}
        if not verify_password(admin_pw, existing["password_hash"]):
            updates["password_hash"] = hash_password(admin_pw)
        if existing.get("name") == "المدير":
            updates["name"] = "ادمن"
        if updates:
            await db.users.update_one({"employee_number": admin_emp}, {"$set": updates})

    # Demo technician accounts removed — use admin panel to create users

    if await db.questions.count_documents({}) == 0:
        for cat, lines in DEFAULT_QUESTIONS.items():
            for i, t in enumerate(lines):
                await db.questions.insert_one({"id": str(uuid.uuid4()), "category": cat,
                                               "text": t, "order": i,
                                               "created_at": datetime.now(timezone.utc).isoformat()})
    else:
        # If panels questions are missing, add them
        if await db.questions.count_documents({"category": "panels"}) == 0:
            for i, t in enumerate(DEFAULT_QUESTIONS["panels"]):
                await db.questions.insert_one({"id": str(uuid.uuid4()), "category": "panels",
                                               "text": t, "order": i,
                                               "created_at": datetime.now(timezone.utc).isoformat()})

    if await db.chillers.count_documents({}) == 0:
        for n in ["C-1", "C-2", "C-3"]:
            await db.chillers.insert_one({"id": str(uuid.uuid4()), "number": n, "name": "",
                                          "created_at": datetime.now(timezone.utc).isoformat()})
    if await db.panels.count_documents({}) == 0:
        for n in ["P-1", "P-2", "P-3"]:
            await db.panels.insert_one({"id": str(uuid.uuid4()), "number": n, "name": "",
                                        "created_at": datetime.now(timezone.utc).isoformat()})

    await db.registration_requests.create_index("id", unique=True)
    await db.registration_requests.create_index("employee_number")
    await db.preventive_plans.create_index("id", unique=True)
    await db.preventive_plans.create_index([("scheduled_date", 1), ("machine_id", 1)])
    await db.breakdowns.create_index("id", unique=True)
    await db.breakdowns.create_index("status")
    await db.breakdowns.create_index("created_at")
    await db.breakdowns.create_index("technician_id")

    await db.downtime_reasons.create_index("id", unique=True)
    await db.downtime_reasons.create_index("order")

    await db.mdb_readings.create_index("id", unique=True)
    await db.mdb_readings.create_index("created_at")
    await db.mdb_readings.create_index("technician_id")
    await db.mdb_readings.create_index("panel_number")
    await db.form_permissions.create_index([("scope", 1), ("target", 1)], unique=True)

    # Seed default downtime reasons if none exist
    if await db.downtime_reasons.count_documents({}) == 0:
        default_reasons = [
            {"text": "Hydraulic Failure",              "specialty": "mechanical"},
            {"text": "Electrical Fault",               "specialty": "electrical"},
            {"text": "Mechanical Breakdown",           "specialty": "mechanical"},
            {"text": "Heating / Temperature Issue",    "specialty": None},
            {"text": "Water / Oil Leakage",            "specialty": "mechanical"},
            {"text": "Abnormal Noise or Vibration",    "specialty": "mechanical"},
            {"text": "Machine Not Starting",           "specialty": "electrical"},
            {"text": "Sensor / Control Error",         "specialty": "electrical"},
            {"text": "Mold / Clamp Issue",             "specialty": "mechanical"},
        ]
        now_iso = datetime.now(timezone.utc).isoformat()
        await db.downtime_reasons.insert_many([
            {"id": str(uuid.uuid4()), "text": r["text"], "specialty": r["specialty"],
             "order": i, "created_at": now_iso}
            for i, r in enumerate(default_reasons)
        ])
        logger.info("Seeded 9 default downtime reasons")

    # Migration: set answer_type=yes_no for any questions missing it (including panels)
    await db.questions.update_many(
        {"answer_type": {"$exists": False}},
        {"$set": {"answer_type": "yes_no"}},
    )

    logger.info("Startup complete")


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)

_cors_origins_env = os.environ.get("CORS_ORIGINS", "")
_cors_origins = (
    [o.strip() for o in _cors_origins_env.split(",") if o.strip()]
    if _cors_origins_env
    else ["https://www.inspet.pro", "https://inspet.pro"]
)
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        return response

app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(GZipMiddleware, minimum_size=1000)
app.add_middleware(CORSMiddleware, allow_credentials=True,
                   allow_origins=_cors_origins,
                   allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
                   allow_headers=["Content-Type", "Authorization"])



